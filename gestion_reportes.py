# gestion_reportes.py
import os
import webbrowser
from datetime import datetime
from typing import Optional

import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from colorama import Fore, Style

from database import db_manager, Equipo, registrar_movimiento_sistema
from ui import mostrar_encabezado, mostrar_menu, pausar_pantalla
from gestion_acceso import requiere_permiso

# --- MENÚ PRINCIPAL DE VISUALIZACIÓN ---
@requiere_permiso("ver_inventario")
def menu_ver_inventario(usuario: str):
    """Menú principal para la visualización de inventario y movimientos."""
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        opciones = [
            "Generar Reportes de Inventario en Excel",
            "Ver últimos 20 movimientos",
            "Ver Inventario Actual en Consola",
            "Volver al menú principal"
        ]
        mostrar_menu(opciones, titulo="Módulo de Visualización de Inventario")
        
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        
        if opcion == '1':
            menu_reportes_excel(usuario)
        elif opcion == '2':
            menu_ver_ultimos_movimientos(usuario)
        elif opcion == '3':
            ver_inventario_consola()
        elif opcion == '4':
            break
        else:
            print(Fore.RED + "Opción no válida.")
            pausar_pantalla()

# --- SUBMENÚS Y FUNCIONES DE VISUALIZACIÓN ---

def menu_ver_ultimos_movimientos(usuario: str):
    """Muestra una tabla con los últimos 20 movimientos de inventario del usuario."""
    os.system('cls' if os.name == 'nt' else 'clear')
    movimientos = db_manager.get_last_movimientos_by_user(usuario, limit=20)
    
    mostrar_encabezado("Tus Últimos 20 Movimientos", color=Fore.BLUE)

    if not movimientos:
        print(Fore.YELLOW + "No has registrado movimientos recientemente.".center(80) + Style.RESET_ALL)
    else:
        print(f"{Fore.CYAN}{'FECHA':<17} {'PLACA':<12} {'MARCA':<15} {'ACCIÓN':<30}{Style.RESET_ALL}")
        print(Fore.CYAN + "-" * 74 + Style.RESET_ALL)
        
        for mov in movimientos:
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            accion = mov.get('accion', 'N/A')
            if len(accion) > 28:
                accion = accion[:27] + "..."

            placa = mov.get('equipo_placa', 'N/A')
            marca = mov.get('marca', 'N/A') or 'N/A'

            print(f"{fecha_formateada:<17} {placa:<12} {marca:<15} {accion:<30}")
    
    pausar_pantalla()

def menu_reportes_excel(usuario: str):
    """Muestra el menú para generar los reportes de inventario en Excel."""
    while True:
        mostrar_menu([
            "Reporte de Inventario Actual (Equipos Activos)",
            "Reporte de Equipos Devueltos a Proveedor",
            "Reporte Histórico Completo de Equipos (Log)",
            "Volver"
        ], titulo="Generar Reportes en Excel")
        
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()

        if opcion == '1':
            generar_excel_inventario(usuario)
        elif opcion == '2':
            generar_excel_devueltos_proveedor(usuario)
        elif opcion == '3':
            generar_excel_historico(usuario)
        elif opcion == '4':
            break
        else:
            print(Fore.RED + "Opción no válida.")

@requiere_permiso("generar_reporte")
def generar_excel_inventario(usuario: str) -> None:
    """Genera un reporte Excel con los equipos activos."""
    try:
        inventario = db_manager.get_equipos_activos()
        if not inventario:
            print(Fore.YELLOW + "\nNo hay equipos activos para generar un reporte.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario de Equipos"

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = [
            "FECHA REGISTRO", "PLACA", "TIPO", "MARCA", "MODELO", "SERIAL", "ESTADO", 
            "FECHA ÚLTIMO CAMBIO", "USUARIO ÚLTIMO CAMBIO", "ASIGNADO A", "EMAIL", "ÚLTIMA OBSERVACIÓN"
        ]
        
        column_widths = {'A': 25, 'B': 15, 'C': 25, 'D': 25, 'E': 25, 'F': 30, 'G': 30, 'H': 25, 'I': 25, 'J': 30, 'K': 30, 'L': 80}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border

        colores_estado = {
            "Disponible": "C6EFCE", "Asignado": "FFEB9C", "En préstamo": "DDEBF7",
            "En mantenimiento": "FCE4D6", "Pendiente Devolución a Proveedor": "FFFFCC"
        }

        for row_num, equipo in enumerate(inventario, 2):
            ultimo_movimiento = db_manager.get_last_movimiento_by_placa(equipo['placa'])
            
            fecha_ult_cambio = "N/A"
            usuario_ult_cambio = "N/A"
            ultima_observacion = equipo.get('observaciones', 'N/A')

            if ultimo_movimiento:
                fecha_obj = datetime.strptime(ultimo_movimiento['fecha'], "%Y-%m-%d %H:%M:%S")
                fecha_ult_cambio = fecha_obj.strftime("%d/%m/%Y %H:%M")
                usuario_ult_cambio = ultimo_movimiento.get('usuario', 'N/A')
                ultima_observacion = ultimo_movimiento.get('detalles', ultima_observacion)

            data_row = [
                equipo.get('fecha_registro', 'N/A'), equipo.get('placa', 'N/A'), equipo.get('tipo', 'N/A'),
                equipo.get('marca', 'N/A'), equipo.get('modelo', 'N/A'), equipo.get('serial', 'N/A'),
                equipo.get('estado', 'N/A'), fecha_ult_cambio, usuario_ult_cambio,
                equipo.get('asignado_a', ''), equipo.get('email_asignado', ''), ultima_observacion
            ]
            
            for col_num, cell_value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
            
            estado_celda = ws.cell(row=row_num, column=7)
            color_hex = colores_estado.get(equipo.get('estado'))
            if color_hex:
                estado_celda.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
        ws.freeze_panes = "A2"
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Inventario Activo", f"Generado reporte con {len(inventario)} equipos", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte de inventario activo en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte Excel: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

@requiere_permiso("generar_reporte")
def generar_excel_devueltos_proveedor(usuario: str) -> None:
    try:
        inventario_devuelto = db_manager.get_equipos_devueltos()
        if not inventario_devuelto:
            print(Fore.YELLOW + "\nNo hay equipos devueltos al proveedor para reportar.")
            pausar_pantalla()
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipos Devueltos"
        
        header_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = [
            "PLACA", "TIPO", "MARCA", "MODELO", "SERIAL", 
            "FECHA DEVOLUCIÓN", "MOTIVO DEVOLUCIÓN", "ÚLTIMA OBSERVACIÓN"
        ]
        
        column_widths = {'A': 15, 'B': 25, 'C': 25, 'D': 25, 'E': 30, 'F': 25, 'G': 25, 'H': 80}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border

        for row_num, equipo in enumerate(inventario_devuelto, 2):
            data_row = [
                equipo.get('placa', 'N/A'), equipo.get('tipo', 'N/A'), equipo.get('marca', 'N/A'),
                equipo.get('modelo', 'N/A'), equipo.get('serial', 'N/A'),
                equipo.get('fecha_devolucion_proveedor', 'N/A'), equipo.get('motivo_devolucion', 'N/A'),
                equipo.get('observaciones', 'N/A')
            ]
            for col_num, cell_value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name
        
        registrar_movimiento_sistema("Reporte Equipos Devueltos", f"Generado reporte con {len(inventario_devuelto)} equipos devueltos", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte de equipos devueltos en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte de equipos devueltos: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

@requiere_permiso("ver_historico")
def generar_excel_historico(usuario: str):
    try:
        log_equipos = db_manager.get_all_log_inventario()

        if not log_equipos:
            print(Fore.YELLOW + "\nNo hay movimientos de equipos para exportar.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico de Movimientos"

        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = ["FECHA", "PLACA EQUIPO", "ACCIÓN", "USUARIO", "DETALLES"]
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 80

        for row_num, mov in enumerate(log_equipos, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('equipo_placa', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=5, value=mov.get('detalles', '')).border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Histórico Equipos", "Generado reporte de histórico de equipos", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte histórico de equipos en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte de histórico: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

@requiere_permiso("ver_inventario")
def ver_inventario_consola():
    mostrar_encabezado("Inventario Actual de Equipos Activos")
    inventario = db_manager.get_equipos_activos()
    if not inventario:
        print(Fore.YELLOW + "\nEl inventario activo está vacío.")
    else:
        print(f"{Fore.CYAN}{'PLACA':<12} {'TIPO':<15} {'MARCA':<15} {'MODELO':<20} {'ESTADO':<30} {'ASIGNADO A':<20}{Style.RESET_ALL}")
        print(Fore.CYAN + "="*112 + Style.RESET_ALL)
        for equipo in inventario:
            estado_color = Fore.WHITE
            if equipo['estado'] == "Disponible": estado_color = Fore.GREEN
            elif equipo['estado'] in ["Asignado", "En préstamo"]: estado_color = Fore.YELLOW
            elif equipo['estado'] == "En mantenimiento": estado_color = Fore.MAGENTA
            elif equipo['estado'] == "Pendiente Devolución a Proveedor": estado_color = Fore.LIGHTYELLOW_EX
            asignado_a = equipo.get('asignado_a') or 'N/A'
            print(f"{equipo['placa']:<12} {equipo['tipo']:<15} {equipo['marca']:<15} {equipo['modelo']:<20} {estado_color}{equipo['estado']:<30}{Style.RESET_ALL} {asignado_a:<20}")
    pausar_pantalla()

def generar_excel_historico_equipo(usuario: str, equipo: Equipo):
    """Genera un reporte Excel con el historial de un solo equipo."""
    try:
        log_equipo = db_manager.get_log_by_placa(equipo.placa)

        if not log_equipo:
            print(Fore.YELLOW + f"\nNo hay historial para el equipo {equipo.placa}.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = f"Historial {equipo.placa}"

        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = ["FECHA", "ACCIÓN", "USUARIO", "DETALLES"]
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 80

        for row_num, mov in enumerate(log_equipo, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('detalles', '')).border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Histórico Individual", f"Generado reporte para placa {equipo.placa}", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el historial del equipo {equipo.placa} en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el historial del equipo: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()