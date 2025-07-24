# inventario.py
import os
import re
import webbrowser
from datetime import datetime
from typing import List, Dict, Optional, Callable
from functools import wraps
import bcrypt
import getpass
from colorama import init, Fore, Back, Style
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sqlite3
import tempfile # Importar para archivos temporales

# Inicializar colorama
init(autoreset=True)

# --- CONFIGURACIÓN ---
DATABASE_NAME = "inventario.db"
USUARIO_ACTUAL = None # Variable global para el usuario logueado

# --- MODELOS DE DATOS ---
class Equipo:
    def __init__(self, placa: str, tipo: str, marca: str, modelo: str, serial: str,
                 estado: str = "Disponible", asignado_a: Optional[str] = None,
                 email_asignado: Optional[str] = None, observaciones: Optional[str] = None,
                 fecha_registro: Optional[str] = None,
                 fecha_devolucion_prestamo: Optional[str] = None,
                 fecha_devolucion_proveedor: Optional[str] = None,
                 motivo_devolucion: Optional[str] = None):
        self.placa = placa
        self.tipo = tipo
        self.marca = marca
        self.modelo = modelo
        self.serial = serial
        self.estado = estado
        self.asignado_a = asignado_a
        self.email_asignado = email_asignado
        self.observaciones = observaciones
        self.fecha_registro = fecha_registro if fecha_registro else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.fecha_devolucion_prestamo = fecha_devolucion_prestamo
        self.fecha_devolucion_proveedor = fecha_devolucion_proveedor
        self.motivo_devolucion = motivo_devolucion

    def to_dict(self) -> Dict:
        return self.__dict__

class MovimientoHistorico:
    def __init__(self, equipo_placa: str, accion: str, detalles: str, usuario: str, fecha: Optional[str] = None):
        self.equipo_placa = equipo_placa
        self.accion = accion
        self.detalles = detalles
        self.usuario = usuario
        self.fecha = fecha if fecha else datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def to_dict(self) -> Dict:
        return self.__dict__

class Usuario:
    def __init__(self, nombre_usuario: str, contrasena_hash: str, rol: str, nombre_completo: Optional[str] = None, cambio_clave_requerido: bool = True, is_active: bool = True):
        self.nombre_usuario = nombre_usuario
        self.contrasena_hash = contrasena_hash
        self.rol = rol
        self.nombre_completo = nombre_completo
        self.cambio_clave_requerido = cambio_clave_requerido
        self.is_active = is_active

    def to_dict(self) -> Dict:
        return self.__dict__

# --- GESTOR DE BASE DE DATOS SQLITE ---
class DatabaseManager:
    def __init__(self, db_name: str):
        self.db_name = db_name
        self.conn = None
        self.connect()
        self.create_tables()
        self.add_missing_columns()

    def connect(self):
        try:
            self.conn = sqlite3.connect(self.db_name)
            self.conn.row_factory = sqlite3.Row
        except sqlite3.Error as e:
            print(Fore.RED + f"❌ Error al conectar a la base de datos: {e}" + Style.RESET_ALL)
            exit()

    def close(self):
        if self.conn:
            self.conn.close()

    def create_tables(self):
        cursor = self.conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS equipos (
                placa TEXT PRIMARY KEY, tipo TEXT NOT NULL, marca TEXT NOT NULL,
                modelo TEXT NOT NULL, serial TEXT NOT NULL, estado TEXT NOT NULL,
                asignado_a TEXT, email_asignado TEXT, observaciones TEXT,
                fecha_registro TEXT, fecha_devolucion_prestamo TEXT, 
                fecha_devolucion_proveedor TEXT, motivo_devolucion TEXT
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS historico (
                id INTEGER PRIMARY KEY AUTOINCREMENT, equipo_placa TEXT NOT NULL,
                accion TEXT NOT NULL, detalles TEXT NOT NULL, usuario TEXT NOT NULL, fecha TEXT NOT NULL
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                nombre_usuario TEXT PRIMARY KEY, contrasena_hash TEXT NOT NULL,
                rol TEXT NOT NULL, nombre_completo TEXT, 
                cambio_clave_requerido INTEGER NOT NULL DEFAULT 1,
                is_active INTEGER NOT NULL DEFAULT 1
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS parametros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tipo TEXT NOT NULL,
                valor TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                UNIQUE(tipo, valor)
            )
        ''')
        self.conn.commit()

    def add_missing_columns(self):
        columns_to_add = {
            'equipos': [
                ('fecha_devolucion_prestamo', 'TEXT'),
                ('fecha_devolucion_proveedor', 'TEXT'),
                ('motivo_devolucion', 'TEXT')
            ],
            'usuarios': [
                ('nombre_completo', 'TEXT'),
                ('cambio_clave_requerido', 'INTEGER NOT NULL DEFAULT 1'),
                ('is_active', 'INTEGER NOT NULL DEFAULT 1')
            ],
            'parametros': [
                ('is_active', 'INTEGER NOT NULL DEFAULT 1')
            ]
        }
        cursor = self.conn.cursor()
        for table, cols in columns_to_add.items():
            for col_name, col_type in cols:
                try:
                    cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type}")
                    self.conn.commit()
                except sqlite3.OperationalError:
                    pass

    def execute_query(self, query: str, params: tuple = ()) -> sqlite3.Cursor:
        cursor = self.conn.cursor()
        cursor.execute(query, params)
        return cursor

    def commit(self):
        self.conn.commit()

    # --- Métodos para Equipos ---
    def insert_equipo(self, equipo: Equipo):
        self.execute_query('''
            INSERT INTO equipos (placa, tipo, marca, modelo, serial, estado, asignado_a, email_asignado, observaciones, fecha_registro, fecha_devolucion_prestamo, fecha_devolucion_proveedor, motivo_devolucion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (equipo.placa, equipo.tipo, equipo.marca, equipo.modelo, equipo.serial, equipo.estado, equipo.asignado_a, equipo.email_asignado, equipo.observaciones, equipo.fecha_registro, equipo.fecha_devolucion_prestamo, equipo.fecha_devolucion_proveedor, equipo.motivo_devolucion))
        self.commit()

    def get_all_equipos(self) -> List[Dict]:
        cursor = self.execute_query('SELECT * FROM equipos')
        return [dict(row) for row in cursor.fetchall()]

    def get_equipo_by_placa(self, placa: str) -> Optional[Dict]:
        cursor = self.execute_query('SELECT * FROM equipos WHERE placa = ?', (placa,))
        row = cursor.fetchone()
        return dict(row) if row else None

    def update_equipo(self, equipo: Equipo):
        self.execute_query('''
            UPDATE equipos SET tipo = ?, marca = ?, modelo = ?, serial = ?, estado = ?, asignado_a = ?,
            email_asignado = ?, observaciones = ?, fecha_devolucion_prestamo = ?, fecha_devolucion_proveedor = ?,
            motivo_devolucion = ? WHERE placa = ?
        ''', (equipo.tipo, equipo.marca, equipo.modelo, equipo.serial, equipo.estado, equipo.asignado_a,
              equipo.email_asignado, equipo.observaciones, equipo.fecha_devolucion_prestamo,
              equipo.fecha_devolucion_proveedor, equipo.motivo_devolucion, equipo.placa))
        self.commit()

    def delete_equipo(self, placa: str):
        self.execute_query('DELETE FROM equipos WHERE placa = ?', (placa,))
        self.commit()

    # --- Métodos para Histórico ---
    def insert_movimiento(self, movimiento: MovimientoHistorico):
        self.execute_query('''
            INSERT INTO historico (equipo_placa, accion, detalles, usuario, fecha) VALUES (?, ?, ?, ?, ?)
        ''', (movimiento.equipo_placa, movimiento.accion, movimiento.detalles, movimiento.usuario, movimiento.fecha))
        self.commit()

    def get_all_historico(self) -> List[Dict]:
        cursor = self.execute_query('SELECT * FROM historico ORDER BY fecha DESC')
        return [dict(row) for row in cursor.fetchall()]

    # --- Métodos para Usuarios ---
    def insert_user(self, user: Usuario):
        self.execute_query('''
            INSERT INTO usuarios (nombre_usuario, contrasena_hash, rol, nombre_completo, cambio_clave_requerido, is_active)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (user.nombre_usuario, user.contrasena_hash, user.rol, user.nombre_completo, int(user.cambio_clave_requerido), int(user.is_active)))
        self.commit()

    def get_user_by_username(self, nombre_usuario: str) -> Optional[Dict]:
        cursor = self.execute_query('SELECT * FROM usuarios WHERE nombre_usuario = ?', (nombre_usuario,))
        row = cursor.fetchone()
        return dict(row) if row else None

    def update_user(self, user: Usuario):
        self.execute_query('''
            UPDATE usuarios SET contrasena_hash = ?, rol = ?, nombre_completo = ?, cambio_clave_requerido = ?, is_active = ?
            WHERE nombre_usuario = ?
        ''', (user.contrasena_hash, user.rol, user.nombre_completo, int(user.cambio_clave_requerido), int(user.is_active), user.nombre_usuario))
        self.commit()

    def get_all_users(self) -> List[Dict]:
        cursor = self.execute_query('SELECT nombre_usuario, rol, nombre_completo, cambio_clave_requerido, is_active FROM usuarios')
        return [dict(row) for row in cursor.fetchall()]

    # --- Métodos para Parámetros ---
    def add_parametro(self, tipo: str, valor: str):
        self.execute_query('INSERT INTO parametros (tipo, valor, is_active) VALUES (?, ?, 1)', (tipo, valor))
        self.commit()

    def get_parametros_por_tipo(self, tipo: str, solo_activos: bool = False) -> List[Dict]:
        query = 'SELECT valor, is_active FROM parametros WHERE tipo = ?'
        if solo_activos:
            query += ' AND is_active = 1'
        query += ' ORDER BY valor'
        cursor = self.execute_query(query, (tipo,))
        return [dict(row) for row in cursor.fetchall()]

    def update_parametro_status(self, tipo: str, valor: str, new_status: bool):
        self.execute_query('UPDATE parametros SET is_active = ? WHERE tipo = ? AND valor = ?', (int(new_status), tipo, valor))
        self.commit()

    def is_parametro_in_use(self, tipo_parametro: str, valor: str) -> bool:
        """
        Verifica si un parámetro está actualmente en uso en la tabla de equipos.
        Retorna True si está en uso, False en caso contrario.
        """
        # Mapea el tipo de parámetro a la columna real en la tabla 'equipos'
        # ej. 'tipo_equipo' -> 'tipo', 'marca_equipo' -> 'marca'
        columna_equipo = tipo_parametro.split('_')[0]

        # Lista blanca de columnas para evitar inyecciones no deseadas
        if columna_equipo not in ['tipo', 'marca']:
            return False

        query = f"SELECT 1 FROM equipos WHERE {columna_equipo} = ? LIMIT 1"
        cursor = self.execute_query(query, (valor,))
        return cursor.fetchone() is not None


db_manager = DatabaseManager(DATABASE_NAME)

# --- FUNCIONES DE PERSISTENCIA ---
def registrar_movimiento(placa: str, accion: str, detalles: str, usuario: str):
    movimiento = MovimientoHistorico(placa, accion, detalles, usuario)
    db_manager.insert_movimiento(movimiento)

# --- FUNCIONES DE UTILIDAD Y VALIDACIÓN ---
def validar_placa_unica(placa: str) -> bool:
    return db_manager.get_equipo_by_placa(placa) is None

def validar_email(email: str) -> bool:
    return re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email) is not None

def validar_placa_formato(placa: str) -> bool:
    return len(placa) >= 4 and placa.isalnum()

def validar_contrasena(contrasena: str) -> bool:
    return len(contrasena) >= 8

def validar_formato_fecha(fecha_str: str) -> Optional[datetime]:
    try:
        return datetime.strptime(fecha_str, "%d/%m/%Y")
    except ValueError:
        return None

# --- FUNCIONES DE VISUALIZACIÓN (UI) ---
def mostrar_encabezado(titulo: str, ancho: int = 80, color: str = Fore.CYAN):
    print("\n" + color + "═" * ancho)
    print(f" {titulo.upper()} ".center(ancho, ' '))
    print("═" * ancho + Style.RESET_ALL)

def mostrar_menu(opciones: List[str], titulo: str):
    mostrar_encabezado(titulo, color=Fore.MAGENTA)
    for i, opcion in enumerate(opciones, 1):
        print(Fore.YELLOW + f"{i}." + Style.RESET_ALL + f" {opcion}")
    print(Fore.MAGENTA + "═" * 80 + Style.RESET_ALL)

def pausar_pantalla():
    input(Fore.CYAN + "\nPresione Enter para continuar..." + Style.RESET_ALL)

# --- CONTROL DE ACCESO BASADO EN ROLES (RBAC) ---
ROLES_PERMISOS = {
    "Administrador": {
        "registrar_equipo", "ver_inventario", "gestionar_equipo", "ver_historico",
        "generar_reporte", "gestionar_usuarios", "eliminar_equipo",
        "devolver_a_proveedor", "aprobar_devoluciones", "gestionar_pendientes",
        "configurar_sistema" # Nuevo permiso
    },
    "Gestor": {
        "registrar_equipo", "ver_inventario", "gestionar_equipo", "ver_historico",
        "generar_reporte", "devolver_a_proveedor"
    },
    "Visualizador": {
        "ver_inventario", "ver_historico", "generar_reporte"
    }
}

def requiere_permiso(permiso: str) -> Callable:
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            global USUARIO_ACTUAL
            if USUARIO_ACTUAL is None:
                print(Fore.RED + "\n❌ Acceso denegado. No hay usuario logueado." + Style.RESET_ALL)
                return
            user_data = db_manager.get_user_by_username(USUARIO_ACTUAL)
            if not user_data:
                print(Fore.RED + "\n❌ Acceso denegado. Usuario no encontrado." + Style.RESET_ALL)
                return
            rol_usuario = user_data['rol']
            if permiso in ROLES_PERMISOS.get(rol_usuario, {}):
                return func(*args, **kwargs)
            else:
                print(Fore.RED + f"\n❌ Permiso denegado. Su rol '{rol_usuario}' no tiene el permiso '{permiso}'." + Style.RESET_ALL)
                pausar_pantalla()
                return
        return wrapper
    return decorator

# --- FUNCIONES DE REPORTES (Excel) ---
@requiere_permiso("generar_reporte")
def generar_excel_inventario(usuario: str, filtro: Optional[str] = None, valor_filtro: Optional[str] = None) -> None:
    try:
        inventario = db_manager.get_all_equipos()

        if not inventario:
            print(Fore.YELLOW + "\nNo hay equipos para generar un reporte.")
            pausar_pantalla()
            return

        datos_a_exportar = inventario
        if filtro and valor_filtro:
            datos_a_exportar = [e for e in inventario if str(e.get(filtro, '')).lower() == valor_filtro.lower()]
        
        if not datos_a_exportar:
            print(Fore.YELLOW + f"\nNo se encontraron equipos con el filtro: {filtro} = '{valor_filtro}'.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario de Equipos"

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = ["PLACA", "TIPO", "MARCA", "MODELO", "SERIAL", "ESTADO", "ASIGNADO A", "EMAIL", "FECHA REGISTRO", "FECHA DEVOLUCIÓN (Préstamo)", "FECHA DEVOLUCIÓN (Proveedor)", "MOTIVO DEVOLUCIÓN", "OBSERVACIONES"]
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border
            ws.column_dimensions[col_letra].width = 22

        colores_estado = {
            "Disponible": "C6EFCE", "Asignado": "FFEB9C", "En préstamo": "DDEBF7",
            "En mantenimiento": "FCE4D6", "Dado de baja": "FFC7CE",
            "Pendiente Devolución a Proveedor": "FFFFCC", "Devuelto a Proveedor": "CCE0B4"
        }

        for row_num, equipo in enumerate(datos_a_exportar, 2):
            ws.cell(row=row_num, column=1, value=equipo.get('placa', 'N/A')).border = border
            ws.cell(row=row_num, column=2, value=equipo.get('tipo', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=equipo.get('marca', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=equipo.get('modelo', 'N/A')).border = border
            ws.cell(row=row_num, column=5, value=equipo.get('serial', 'N/A')).border = border
            
            estado_celda = ws.cell(row=row_num, column=6, value=equipo.get('estado', 'N/A'))
            estado_celda.border = border
            color_hex = colores_estado.get(equipo.get('estado'))
            if color_hex:
                estado_celda.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

            ws.cell(row=row_num, column=7, value=equipo.get('asignado_a', '')).border = border
            ws.cell(row=row_num, column=8, value=equipo.get('email_asignado', '')).border = border
            ws.cell(row=row_num, column=9, value=equipo.get('fecha_registro', '')).border = border
            ws.cell(row=row_num, column=10, value=equipo.get('fecha_devolucion_prestamo', '')).border = border
            ws.cell(row=row_num, column=11, value=equipo.get('fecha_devolucion_proveedor', '')).border = border
            ws.cell(row=row_num, column=12, value=equipo.get('motivo_devolucion', '')).border = border
            ws.cell(row=row_num, column=13, value=equipo.get('observaciones', '')).border = border
        
        ws.freeze_panes = "A2"
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento("SISTEMA", "Reporte Excel", f"Generado reporte de inventario con {len(datos_a_exportar)} equipos", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte de inventario en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte Excel: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

@requiere_permiso("ver_historico")
def generar_excel_historico(usuario: str):
    try:
        historico = db_manager.get_all_historico()

        if not historico:
            print(Fore.YELLOW + "\nNo hay movimientos históricos para exportar.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico de Movimientos"

        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = ["FECHA", "PLACA EQUIPO", "ACCIÓN", "USUARIO", "DETALLES"]
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border
        
        ws.column_dimensions['A'].width = 25 # Fecha
        ws.column_dimensions['B'].width = 20 # Placa
        ws.column_dimensions['C'].width = 25 # Acción
        ws.column_dimensions['D'].width = 20 # Usuario
        ws.column_dimensions['E'].width = 80 # Detalles

        for row_num, mov in enumerate(historico, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('equipo_placa', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=5, value=mov.get('detalles', '')).border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento("SISTEMA", "Reporte Histórico Excel", "Generado reporte de histórico", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte de histórico en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte de histórico: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

# --- FUNCIONES PRINCIPALES DE INVENTARIO ---
def seleccionar_parametro(tipo_parametro: str, nombre_amigable: str) -> str:
    """Función auxiliar para seleccionar un parámetro de una lista o ingresarlo manualmente."""
    parametros_activos = [p['valor'] for p in db_manager.get_parametros_por_tipo(tipo_parametro, solo_activos=True)]
    
    if not parametros_activos:
        print(Fore.YELLOW + f"No hay {nombre_amigable}s preconfigurados. Se solicitará entrada manual.")
        return input(Fore.YELLOW + f"{nombre_amigable} del equipo: " + Style.RESET_ALL).strip()
    
    while True:
        print(Fore.CYAN + f"\nSeleccione un {nombre_amigable}:")
        for i, param in enumerate(parametros_activos, 1):
            print(f"{i}. {param}")
        
        seleccion = input(Fore.YELLOW + "Opción: " + Style.RESET_ALL).strip()
        try:
            idx = int(seleccion) - 1
            if 0 <= idx < len(parametros_activos):
                return parametros_activos[idx]
            else:
                print(Fore.RED + "Selección fuera de rango.")
        except ValueError:
            print(Fore.RED + "Por favor, ingrese un número.")

@requiere_permiso("registrar_equipo")
def registrar_equipo(usuario: str):
    mostrar_encabezado("Registro de Nuevo Equipo", color=Fore.BLUE)
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        while True:
            placa = input(Fore.YELLOW + "Placa del equipo: " + Style.RESET_ALL).strip().upper()
            if not validar_placa_formato(placa):
                print(Fore.RED + "⚠️ Formato de placa inválido (mín. 4 caracteres alfanuméricos).")
            elif not validar_placa_unica(placa):
                print(Fore.RED + "⚠️ Esta placa ya está registrada.")
            else:
                break
        
        tipo = seleccionar_parametro('tipo_equipo', 'Tipo de Equipo')
        marca = seleccionar_parametro('marca_equipo', 'Marca')
        modelo = input(Fore.YELLOW + "Modelo: " + Style.RESET_ALL).strip()
        serial = input(Fore.YELLOW + "Número de serie: " + Style.RESET_ALL).strip()
        observaciones = input(Fore.YELLOW + "Observaciones (opcional): " + Style.RESET_ALL).strip() or None

        if not all([placa, tipo, marca, modelo, serial]):
            print(Fore.RED + "❌ Error: Todos los campos son obligatorios excepto observaciones.")
            return

        nuevo_equipo = Equipo(placa=placa, tipo=tipo, marca=marca, modelo=modelo, serial=serial, observaciones=observaciones)
        db_manager.insert_equipo(nuevo_equipo)
        registrar_movimiento(placa, "Registro", f"Nuevo equipo registrado: {tipo} {marca} {modelo}", usuario)
        print(Fore.GREEN + f"\n✅ Equipo con placa {placa} registrado exitosamente!")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación de registro cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def gestionar_equipo_por_placa(usuario: str):
    mostrar_encabezado("Gestión de Equipo por Placa", color=Fore.BLUE)
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        placa = input(Fore.YELLOW + "Ingrese la placa del equipo a gestionar: " + Style.RESET_ALL).strip().upper()
        equipo_data = db_manager.get_equipo_by_placa(placa)

        if not equipo_data:
            print(Fore.RED + "❌ No se encontró un equipo con esa placa.")
            pausar_pantalla()
            return
        
        equipo = Equipo(**equipo_data)
        menu_gestion_especifica(usuario, equipo)

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación de gestión cancelada.")

def menu_gestion_especifica(usuario: str, equipo: Equipo):
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        mostrar_encabezado(f"Gestionando Equipo: {equipo.marca} {equipo.modelo} - PLACA: {equipo.placa}", color=Fore.GREEN)
        print(f"{Fore.CYAN}Estado actual:{Style.RESET_ALL} {equipo.estado}")
        if equipo.asignado_a: print(f"{Fore.CYAN}Asignado a:{Style.RESET_ALL} {equipo.asignado_a} ({equipo.email_asignado})")
        if equipo.fecha_devolucion_prestamo: print(f"{Fore.CYAN}Fecha devolución (Préstamo):{Style.RESET_ALL} {equipo.fecha_devolucion_prestamo}")
        if equipo.fecha_devolucion_proveedor: print(f"{Fore.CYAN}Fecha devolución (Proveedor):{Style.RESET_ALL} {equipo.fecha_devolucion_proveedor}")
        if equipo.motivo_devolucion: print(f"{Fore.CYAN}Motivo devolución:{Style.RESET_ALL} {equipo.motivo_devolucion}")
        print("-" * 80)

        opciones_gestion = [
            "Asignar/Prestar equipo", "Devolver equipo al inventario", "Registrar para mantenimiento",
            "Registrar para devolución a Proveedor", "Editar información del equipo", 
            "Eliminar equipo", "Volver al menú anterior"
        ]
        mostrar_menu(opciones_gestion, titulo=f"Opciones para {equipo.placa}")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()

        if opcion == "1": asignar_o_prestar_equipo(usuario, equipo)
        elif opcion == "2": devolver_equipo(usuario, equipo)
        elif opcion == "3": registrar_mantenimiento(usuario, equipo)
        elif opcion == "4": registrar_devolucion_a_proveedor(usuario, equipo)
        elif opcion == "5": editar_equipo(usuario, equipo)
        elif opcion == "6":
            if eliminar_equipo(usuario, equipo): return
        elif opcion == "7": break
        else:
            print(Fore.RED + "❌ Opción no válida.")
            pausar_pantalla()
        
        equipo_data = db_manager.get_equipo_by_placa(equipo.placa)
        if not equipo_data: break
        equipo = Equipo(**equipo_data)

@requiere_permiso("gestionar_equipo")
def asignar_o_prestar_equipo(usuario: str, equipo: Equipo):
    if equipo.estado != "Disponible":
        print(Fore.RED + f"❌ El equipo no está 'Disponible' (Estado actual: {equipo.estado}).")
        pausar_pantalla()
        return
    
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        tipo_asignacion = input(Fore.YELLOW + "Escriba 'A' para Asignación o 'P' para Préstamo: " + Style.RESET_ALL).strip().upper()
        if tipo_asignacion not in ["A", "P"]:
            print(Fore.RED + "Opción inválida."); return
        
        nombre_asignado = input(Fore.YELLOW + "Nombre de la persona: " + Style.RESET_ALL).strip()
        while True:
            email_asignado = input(Fore.YELLOW + "Email de la persona: " + Style.RESET_ALL).strip()
            if validar_email(email_asignado): break
            print(Fore.RED + "Email inválido.")

        if tipo_asignacion == "P":
            equipo.estado = "En préstamo"
            while True:
                fecha_str = input(Fore.YELLOW + "Fecha de devolución (DD/MM/AAAA): " + Style.RESET_ALL).strip()
                if validar_formato_fecha(fecha_str):
                    equipo.fecha_devolucion_prestamo = fecha_str
                    break
                print(Fore.RED + "Formato de fecha inválido.")
            detalles_movimiento = f"Préstamo hasta: {equipo.fecha_devolucion_prestamo}"
        else:
            equipo.estado = "Asignado"
            detalles_movimiento = "Asignación permanente"
        
        equipo.asignado_a = nombre_asignado
        equipo.email_asignado = email_asignado
        db_manager.update_equipo(equipo)
        registrar_movimiento(equipo.placa, "Asignación/Préstamo", f"{detalles_movimiento} a: {nombre_asignado}", usuario)
        print(Fore.GREEN + f"\n✅ Equipo {equipo.placa} {equipo.estado.lower()} a {nombre_asignado}.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def devolver_equipo(usuario: str, equipo: Equipo):
    if equipo.estado not in ["Asignado", "En préstamo"]:
        print(Fore.RED + "❌ El equipo no está asignado ni en préstamo.")
        pausar_pantalla()
        return

    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        detalles_previos = f"Estado anterior: {equipo.estado}. Asignado a: {equipo.asignado_a or 'N/A'}"
        equipo.estado = "Disponible"
        equipo.asignado_a = None
        equipo.email_asignado = None
        equipo.fecha_devolucion_prestamo = None
        db_manager.update_equipo(equipo)
        registrar_movimiento(equipo.placa, "Devolución a Inventario", f"Equipo devuelto. {detalles_previos}", usuario)
        print(Fore.GREEN + f"\n✅ Equipo {equipo.placa} marcado como 'Disponible'.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def registrar_mantenimiento(usuario: str, equipo: Equipo):
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        observaciones_mantenimiento = input(Fore.YELLOW + "Observaciones del mantenimiento: " + Style.RESET_ALL).strip()
        if not observaciones_mantenimiento:
            print(Fore.RED + "Las observaciones son obligatorias."); return
        
        estado_anterior = equipo.estado
        equipo.estado = "En mantenimiento"
        db_manager.update_equipo(equipo)
        registrar_movimiento(equipo.placa, "Mantenimiento", f"Observaciones: {observaciones_mantenimiento}. Estado anterior: {estado_anterior}", usuario)
        print(Fore.GREEN + f"\n✅ Mantenimiento registrado. Estado cambiado a 'En mantenimiento'.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("devolver_a_proveedor")
def registrar_devolucion_a_proveedor(usuario: str, equipo: Equipo):
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        while True:
            print("Motivo de la devolución:\n1. Por daño\n2. No se necesita más")
            motivo_opcion = input(Fore.YELLOW + "Seleccione el motivo: " + Style.RESET_ALL).strip()
            if motivo_opcion == '1':
                motivo = "Por daño"
                break
            elif motivo_opcion == '2':
                motivo = "No se necesita más"
                break
            else:
                print(Fore.RED + "Opción inválida.")

        while True:
            fecha_str = input(Fore.YELLOW + "Fecha de devolución a proveedor (DD/MM/AAAA): " + Style.RESET_ALL).strip()
            if validar_formato_fecha(fecha_str):
                fecha_devolucion = fecha_str
                break
            print(Fore.RED + "Formato de fecha inválido.")

        observaciones = input(Fore.YELLOW + "Observaciones adicionales: " + Style.RESET_ALL).strip()
        
        estado_anterior = equipo.estado
        equipo.estado = "Pendiente Devolución a Proveedor"
        equipo.fecha_devolucion_proveedor = fecha_devolucion
        equipo.motivo_devolucion = motivo
        equipo.observaciones = observaciones
        equipo.asignado_a = None
        equipo.email_asignado = None
        equipo.fecha_devolucion_prestamo = None

        db_manager.update_equipo(equipo)
        detalles = f"Motivo: {motivo}. Fecha prog.: {fecha_devolucion}. Obs: {observaciones}. Estado anterior: {estado_anterior}"
        registrar_movimiento(equipo.placa, "Registro Devolución Proveedor", detalles, usuario)
        print(Fore.GREEN + f"\n✅ Equipo {equipo.placa} registrado para devolución a proveedor.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def editar_equipo(usuario: str, equipo: Equipo):
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        print(Fore.CYAN + "Deje en blanco para mantener el valor actual.")
        tipo_nuevo = input(Fore.YELLOW + f"Tipo ({equipo.tipo}): " + Style.RESET_ALL).strip() or equipo.tipo
        marca_nueva = input(Fore.YELLOW + f"Marca ({equipo.marca}): " + Style.RESET_ALL).strip() or equipo.marca
        modelo_nuevo = input(Fore.YELLOW + f"Modelo ({equipo.modelo}): " + Style.RESET_ALL).strip() or equipo.modelo
        serial_nuevo = input(Fore.YELLOW + f"Serie ({equipo.serial}): " + Style.RESET_ALL).strip() or equipo.serial
        observaciones_nuevas = input(Fore.YELLOW + f"Observaciones ({equipo.observaciones or ''}): " + Style.RESET_ALL).strip() or equipo.observaciones

        cambios = []
        if equipo.tipo != tipo_nuevo: cambios.append(f"Tipo: '{equipo.tipo}' -> '{tipo_nuevo}'")
        if equipo.marca != marca_nueva: cambios.append(f"Marca: '{equipo.marca}' -> '{marca_nueva}'")
        if equipo.modelo != modelo_nuevo: cambios.append(f"Modelo: '{equipo.modelo}' -> '{modelo_nuevo}'")
        if equipo.serial != serial_nuevo: cambios.append(f"Serial: '{equipo.serial}' -> '{serial_nuevo}'")
        if equipo.observaciones != observaciones_nuevas: cambios.append(f"Observaciones: '{equipo.observaciones or ''}' -> '{observaciones_nuevas}'")
        
        if not cambios:
            print(Fore.YELLOW + "\nNo se detectaron cambios."); return

        equipo.tipo, equipo.marca, equipo.modelo, equipo.serial, equipo.observaciones = tipo_nuevo, marca_nueva, modelo_nuevo, serial_nuevo, observaciones_nuevas
        db_manager.update_equipo(equipo)
        registrar_movimiento(equipo.placa, "Edición", "; ".join(cambios), usuario)
        print(Fore.GREEN + f"\n✅ Equipo {equipo.placa} actualizado.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("eliminar_equipo")
def eliminar_equipo(usuario: str, equipo: Equipo) -> bool:
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        confirmacion = input(Fore.RED + f"⚠️ ¿Seguro de eliminar el equipo {equipo.placa}? Esta acción es irreversible. (Escriba 'SI'): " + Style.RESET_ALL).strip().upper()
        if confirmacion == "SI":
            db_manager.delete_equipo(equipo.placa)
            registrar_movimiento(equipo.placa, "Eliminación", f"Equipo eliminado: {equipo.tipo} {equipo.marca}", usuario)
            print(Fore.GREEN + f"\n✅ Equipo {equipo.placa} eliminado.")
            pausar_pantalla()
            return True
        else:
            print(Fore.YELLOW + "\nOperación de eliminación cancelada.")
            pausar_pantalla()
            return False
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
        pausar_pantalla()
        return False

# --- MÓDULOS DE GESTIÓN ADMINISTRATIVA ---
@requiere_permiso("gestionar_pendientes")
def menu_gestionar_pendientes(usuario: str):
    while True:
        mostrar_menu([
            "Gestionar Equipos en Mantenimiento",
            "Gestionar Devoluciones a Proveedor Pendientes",
            "Volver"
        ], titulo="Gestionar Mantenimientos y Devoluciones")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        if opcion == '1':
            gestionar_mantenimientos(usuario)
        elif opcion == '2':
            gestionar_devoluciones_proveedor(usuario)
        elif opcion == '3':
            break
        else:
            print(Fore.RED + "Opción no válida.")

def gestionar_mantenimientos(usuario: str):
    mostrar_encabezado("Gestionar Equipos en Mantenimiento", color=Fore.BLUE)
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        while True:
            equipos_pendientes = [Equipo(**e) for e in db_manager.get_all_equipos() if e.get('estado') == "En mantenimiento"]
            if not equipos_pendientes:
                print(Fore.YELLOW + "\nNo hay equipos en mantenimiento para gestionar.")
                break

            os.system('cls' if os.name == 'nt' else 'clear')
            mostrar_encabezado("Gestionar Equipos en Mantenimiento", color=Fore.BLUE)
            print(Fore.WHITE + "\n--- Equipos en Mantenimiento ---" + Style.RESET_ALL)
            for i, equipo in enumerate(equipos_pendientes, 1):
                print(f"{Fore.YELLOW}{i}.{Style.RESET_ALL} Placa: {equipo.placa}, Tipo: {equipo.tipo}, Marca: {equipo.marca}")
            print(Fore.WHITE + "---------------------------------" + Style.RESET_ALL)

            seleccion = input(Fore.YELLOW + "Seleccione el equipo a gestionar (o '0' para salir): " + Style.RESET_ALL).strip()
            if seleccion == '0': break
            
            try:
                indice = int(seleccion) - 1
                if not (0 <= indice < len(equipos_pendientes)):
                    print(Fore.RED + "❌ Número no válido."); continue
                
                equipo_a_gestionar = equipos_pendientes[indice]
                print(f"\nGestionando: {equipo_a_gestionar.placa}")
                print("1. Mantenimiento completado (Pasa a 'Disponible')")
                print("2. Equipo no reparable (Registrar para devolución a proveedor)")
                print("0. Cancelar")
                accion = input(Fore.YELLOW + "Seleccione una acción: " + Style.RESET_ALL).strip()

                if accion == '1':
                    estado_anterior = equipo_a_gestionar.estado
                    equipo_a_gestionar.estado = "Disponible"
                    db_manager.update_equipo(equipo_a_gestionar)
                    registrar_movimiento(equipo_a_gestionar.placa, "Mantenimiento Completado", f"Estado cambiado de '{estado_anterior}' a 'Disponible'", usuario)
                    print(Fore.GREEN + f"\n✅ Equipo {equipo_a_gestionar.placa} ahora está 'Disponible'.")
                elif accion == '2':
                    print(Fore.CYAN + "\nRedirigiendo...")
                    pausar_pantalla()
                    registrar_devolucion_a_proveedor(usuario, equipo_a_gestionar)
                elif accion == '0': continue
                else: print(Fore.RED + "❌ Acción no válida.")
            except ValueError:
                print(Fore.RED + "❌ Entrada inválida. Ingrese un número.")
            pausar_pantalla()
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

def gestionar_devoluciones_proveedor(usuario: str):
    mostrar_encabezado("Gestionar Devoluciones a Proveedor", color=Fore.BLUE)
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        while True:
            equipos_pendientes = [Equipo(**e) for e in db_manager.get_all_equipos() if e.get('estado') == "Pendiente Devolución a Proveedor"]
            if not equipos_pendientes:
                print(Fore.YELLOW + "\nNo hay devoluciones pendientes para gestionar.")
                break

            os.system('cls' if os.name == 'nt' else 'clear')
            mostrar_encabezado("Gestionar Devoluciones a Proveedor", color=Fore.BLUE)
            print(Fore.WHITE + "\n--- Devoluciones Pendientes ---" + Style.RESET_ALL)
            for i, equipo in enumerate(equipos_pendientes, 1):
                print(f"{Fore.YELLOW}{i}.{Style.RESET_ALL} Placa: {equipo.placa}, Fecha Prog.: {equipo.fecha_devolucion_proveedor}, Motivo: {equipo.motivo_devolucion}")
            print(Fore.WHITE + "---------------------------------" + Style.RESET_ALL)

            seleccion = input(Fore.YELLOW + "Seleccione el equipo a confirmar como devuelto (o '0' para salir): " + Style.RESET_ALL).strip()
            if seleccion == '0': break
            
            try:
                indice = int(seleccion) - 1
                if not (0 <= indice < len(equipos_pendientes)):
                    print(Fore.RED + "❌ Número no válido."); continue
                
                equipo_a_gestionar = equipos_pendientes[indice]
                confirm = input(Fore.YELLOW + f"¿Confirmar que el equipo {equipo_a_gestionar.placa} ha sido devuelto al proveedor? (S/N): " + Style.RESET_ALL).strip().upper()

                if confirm == 'S':
                    estado_anterior = equipo_a_gestionar.estado
                    equipo_a_gestionar.estado = "Devuelto a Proveedor"
                    db_manager.update_equipo(equipo_a_gestionar)
                    registrar_movimiento(equipo_a_gestionar.placa, "Devolución a Proveedor Completada", f"Estado cambiado de '{estado_anterior}' a 'Devuelto a Proveedor'", usuario)
                    print(Fore.GREEN + f"\n✅ Equipo {equipo_a_gestionar.placa} marcado como 'Devuelto a Proveedor'.")
                else:
                    print(Fore.YELLOW + "Operación cancelada.")
            except ValueError:
                print(Fore.RED + "❌ Entrada inválida. Ingrese un número.")
            pausar_pantalla()
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

# --- FUNCIONES DE AUTENTICACIÓN Y GESTIÓN DE USUARIOS ---
def hash_contrasena(contrasena: str) -> str:
    return bcrypt.hashpw(contrasena.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verificar_contrasena(contrasena: str, hash_almacenado: str) -> bool:
    return bcrypt.checkpw(contrasena.encode('utf-8'), hash_almacenado.encode('utf-8'))

def login() -> Optional[str]:
    global USUARIO_ACTUAL
    mostrar_encabezado("Inicio de Sesión", color=Fore.GREEN)
    for _ in range(3):
        nombre_usuario = input(Fore.YELLOW + "Usuario: " + Style.RESET_ALL).strip()
        contrasena = getpass.getpass(Fore.YELLOW + "Contraseña: " + Style.RESET_ALL)
        user_data = db_manager.get_user_by_username(nombre_usuario)
        if user_data and verificar_contrasena(contrasena, user_data['contrasena_hash']):
            if not user_data['is_active']:
                print(Fore.RED + "❌ Su cuenta de usuario está bloqueada. Contacte a un administrador.")
                continue
            
            USUARIO_ACTUAL = nombre_usuario
            print(Fore.GREEN + f"\n✅ ¡Bienvenido, {user_data.get('nombre_completo', nombre_usuario)}!")
            if user_data.get('cambio_clave_requerido'):
                print(Fore.YELLOW + "⚠️ Su contraseña debe ser cambiada.")
                cambiar_contrasena_usuario(nombre_usuario, forzar_cambio=True)
            pausar_pantalla()
            return nombre_usuario
        else:
            print(Fore.RED + "❌ Credenciales incorrectas.")
    print(Fore.RED + "\n❌ Demasiados intentos fallidos.")
    return None

@requiere_permiso("gestionar_usuarios")
def registrar_usuario(usuario_actual: str):
    mostrar_encabezado("Registrar Nuevo Usuario")
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        nombre_completo = input(Fore.YELLOW + "Nombre completo: " + Style.RESET_ALL).strip()
        nombre_usuario = input(Fore.YELLOW + "Nombre de usuario: " + Style.RESET_ALL).strip().lower()
        if not nombre_completo or not nombre_usuario:
            print(Fore.RED + "Nombre completo y nombre de usuario son obligatorios."); return
        if db_manager.get_user_by_username(nombre_usuario):
            print(Fore.RED + "Este nombre de usuario ya existe."); return
        
        contrasena = getpass.getpass(Fore.YELLOW + "Contraseña (mín. 8 caracteres): " + Style.RESET_ALL)
        if not validar_contrasena(contrasena):
            print(Fore.RED + "La contraseña es muy corta."); return

        roles_validos = ["Administrador", "Gestor", "Visualizador"]
        rol = input(Fore.YELLOW + f"Rol ({', '.join(roles_validos)}): " + Style.RESET_ALL).strip().capitalize()
        if rol not in roles_validos:
            print(Fore.RED + "Rol no válido."); return

        nuevo_usuario = Usuario(nombre_usuario, hash_contrasena(contrasena), rol, nombre_completo, True, True)
        db_manager.insert_user(nuevo_usuario)
        registrar_movimiento("SISTEMA", "Registro Usuario", f"Usuario '{nombre_usuario}' ({rol}) registrado por {usuario_actual}", usuario_actual)
        print(Fore.GREEN + f"\n✅ Usuario '{nombre_usuario}' registrado.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

def cambiar_contrasena_usuario(nombre_usuario: str, forzar_cambio: bool = False):
    mostrar_encabezado(f"Cambiar Contraseña para {nombre_usuario}")
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    user_data = db_manager.get_user_by_username(nombre_usuario)
    if not user_data: print(Fore.RED + "Usuario no encontrado."); return
    try:
        if not forzar_cambio:
            old_password = getpass.getpass(Fore.YELLOW + "Contraseña actual: " + Style.RESET_ALL)
            if not verificar_contrasena(old_password, user_data['contrasena_hash']):
                print(Fore.RED + "Contraseña actual incorrecta."); return
        
        new_password = getpass.getpass(Fore.YELLOW + "Nueva contraseña (mín. 8 caracteres): " + Style.RESET_ALL)
        if not validar_contrasena(new_password):
            print(Fore.RED + "La nueva contraseña es muy corta."); return
        
        user_obj = Usuario(**user_data)
        user_obj.contrasena_hash = hash_contrasena(new_password)
        user_obj.cambio_clave_requerido = False
        db_manager.update_user(user_obj)
        registrar_movimiento("SISTEMA", "Cambio Contraseña", f"Contraseña cambiada para '{nombre_usuario}'", nombre_usuario)
        print(Fore.GREEN + "\n✅ Contraseña cambiada exitosamente.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

def inicializar_admin_si_no_existe():
    if not db_manager.get_user_by_username("admin"):
        print(Fore.YELLOW + "\nCreando usuario administrador inicial 'admin'...")
        admin_pass_hash = hash_contrasena("adminpass")
        admin_user = Usuario("admin", admin_pass_hash, "Administrador", "Administrador Principal", True, True)
        db_manager.insert_user(admin_user)
        print(Fore.GREEN + "✅ Usuario 'admin' creado con contraseña 'adminpass'. Por favor, cámbiela.")
        pausar_pantalla()

# --- MENÚS MODULARES ---
@requiere_permiso("ver_inventario")
def menu_ver_inventario(usuario: str):
    while True:
        mostrar_menu(["Ver en Consola", "Exportar a Excel (Completo)", "Volver"], titulo="Ver Inventario")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        if opcion == '1': ver_inventario_consola()
        elif opcion == '2': generar_excel_inventario(usuario)
        elif opcion == '3': break
        else: print(Fore.RED + "Opción no válida.")

@requiere_permiso("generar_reporte")
def menu_reportes(usuario: str):
    while True:
        mostrar_menu([
            "Reporte completo",
            "Reporte por Estado (ej. Asignado, Disponible)",
            "Volver"
        ], titulo="Generar Reportes")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        if opcion == '1':
            generar_excel_inventario(usuario)
        elif opcion == '2':
            estado = input(Fore.YELLOW + "Ingrese el estado a filtrar: " + Style.RESET_ALL).strip().title()
            generar_excel_inventario(usuario, filtro='estado', valor_filtro=estado)
        elif opcion == '3':
            break
        else:
            print(Fore.RED + "Opción no válida.")

def menu_gestion_inventario(usuario: str):
    user_data = db_manager.get_user_by_username(usuario)
    rol_actual = user_data['rol']
    
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        opciones_disponibles = []
        if "registrar_equipo" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Registrar nuevo equipo")
        if "gestionar_equipo" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Gestionar un equipo por placa")
        if "ver_inventario" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Ver inventario (Consola/Excel)")
        if "generar_reporte" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Generar reportes avanzados")
        if "gestionar_pendientes" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Gestionar Mantenimientos y Devoluciones")
        opciones_disponibles.append("Volver al menú principal")

        mostrar_menu(opciones_disponibles, titulo="Módulo de Gestión de Inventario")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        
        try:
            opcion_idx = int(opcion) - 1
            if 0 <= opcion_idx < len(opciones_disponibles):
                opcion_texto = opciones_disponibles[opcion_idx]
                if opcion_texto == "Registrar nuevo equipo": registrar_equipo(usuario)
                elif opcion_texto == "Gestionar un equipo por placa": gestionar_equipo_por_placa(usuario)
                elif opcion_texto == "Ver inventario (Consola/Excel)": menu_ver_inventario(usuario)
                elif opcion_texto == "Generar reportes avanzados": menu_reportes(usuario)
                elif opcion_texto == "Gestionar Mantenimientos y Devoluciones": menu_gestionar_pendientes(usuario)
                elif opcion_texto == "Volver al menú principal": break
            else: print(Fore.RED + "Opción no válida.")
        except (ValueError, IndexError):
            print(Fore.RED + "Entrada no válida.")

def menu_gestion_acceso_sistema(usuario: str):
    user_data = db_manager.get_user_by_username(usuario)
    rol_actual = user_data['rol']
    
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        opciones_disponibles = []
        if "gestionar_usuarios" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Gestión de usuarios")
        if "ver_historico" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Ver histórico de movimientos")
        if "configurar_sistema" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("Configuración del Sistema")
        opciones_disponibles.append("Cambiar mi contraseña")
        opciones_disponibles.append("Volver al menú principal")
        
        mostrar_menu(opciones_disponibles, titulo="Módulo de Acceso y Sistema")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        
        try:
            opcion_idx = int(opcion) - 1
            if 0 <= opcion_idx < len(opciones_disponibles):
                opcion_texto = opciones_disponibles[opcion_idx]
                if opcion_texto == "Gestión de usuarios": menu_usuarios(usuario)
                elif opcion_texto == "Ver histórico de movimientos": menu_ver_historico(usuario)
                elif opcion_texto == "Configuración del Sistema": menu_configuracion_sistema(usuario)
                elif opcion_texto == "Cambiar mi contraseña": cambiar_contrasena_usuario(usuario)
                elif opcion_texto == "Volver al menú principal": break
            else: print(Fore.RED + "Opción no válida.")
        except (ValueError, IndexError):
            print(Fore.RED + "Entrada no válida.")

@requiere_permiso("gestionar_usuarios")
def menu_usuarios(usuario_actual: str):
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        mostrar_encabezado("Gestión de Usuarios")
        usuarios = db_manager.get_all_users()
        
        print(f"{Fore.CYAN}{'USUARIO':<20} {'NOMBRE COMPLETO':<30} {'ESTADO'}{Style.RESET_ALL}")
        print(Fore.CYAN + "-" * 65 + Style.RESET_ALL)
        if not usuarios:
            print(Fore.YELLOW + "No hay usuarios registrados.")
        else:
            for user in usuarios:
                estado = Fore.GREEN + "Activo" if user['is_active'] else Fore.RED + "Bloqueado"
                nombre_completo = user.get('nombre_completo') or 'N/A'
                print(f"{user['nombre_usuario']:<20} {nombre_completo:<30} {estado}{Style.RESET_ALL}")
        print(Fore.CYAN + "-" * 65 + Style.RESET_ALL)
        
        mostrar_menu(["Registrar nuevo usuario", "Gestionar un usuario existente", "Volver"], titulo="Opciones de Gestión de Usuarios")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()

        if opcion == '1':
            registrar_usuario(usuario_actual)
        elif opcion == '2':
            if not usuarios:
                print(Fore.YELLOW + "No hay usuarios para gestionar.")
                pausar_pantalla()
                continue
            target_username = input(Fore.YELLOW + "Ingrese el nombre de usuario a gestionar: " + Style.RESET_ALL).strip().lower()
            target_user_data = db_manager.get_user_by_username(target_username)
            if target_user_data:
                gestionar_usuario_especifico(usuario_actual, target_user_data)
            else:
                print(Fore.RED + "Usuario no encontrado.")
                pausar_pantalla()
        elif opcion == '3':
            break
        else:
            print(Fore.RED + "Opción no válida.")
            pausar_pantalla()

def gestionar_usuario_especifico(admin_usuario: str, target_user_data: Dict):
    target_user_obj = Usuario(**db_manager.get_user_by_username(target_user_data['nombre_usuario']))
    
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        estado = Fore.GREEN + "Activo" if target_user_obj.is_active else Fore.RED + "Bloqueado"
        mostrar_encabezado(f"Gestionando a {target_user_obj.nombre_usuario}", color=Fore.GREEN)
        
        opcion_bloqueo = (Fore.RED + "Bloquear Acceso") if target_user_obj.is_active else (Fore.GREEN + "Desbloquear Acceso")
        
        opciones_menu = [
            "Modificar nombre completo",
            "Resetear contraseña",
            opcion_bloqueo,
            "Volver"
        ]
        
        titulo_menu = f"OPCIONES - {target_user_obj.nombre_completo or 'N/A'} / {target_user_obj.nombre_usuario}"
        mostrar_menu(opciones_menu, titulo=titulo_menu)
        opcion = input(Fore.YELLOW + "Seleccione: " + Style.RESET_ALL).strip()

        if opcion == '1':
            nuevo_nombre = input(Fore.YELLOW + f"Nuevo nombre completo ({target_user_obj.nombre_completo}): " + Style.RESET_ALL).strip()
            if nuevo_nombre:
                target_user_obj.nombre_completo = nuevo_nombre
                db_manager.update_user(target_user_obj)
                registrar_movimiento("SISTEMA", "Modificación Usuario", f"Nombre de '{target_user_obj.nombre_usuario}' cambiado a '{nuevo_nombre}' por {admin_usuario}", admin_usuario)
                print(Fore.GREEN + "Nombre actualizado.")
        elif opcion == '2':
            nueva_pass = getpass.getpass(Fore.YELLOW + f"Ingrese la nueva contraseña para {target_user_obj.nombre_usuario}: " + Style.RESET_ALL)
            if validar_contrasena(nueva_pass):
                target_user_obj.contrasena_hash = hash_contrasena(nueva_pass)
                target_user_obj.cambio_clave_requerido = True
                db_manager.update_user(target_user_obj)
                registrar_movimiento("SISTEMA", "Reset Contraseña", f"Contraseña de '{target_user_obj.nombre_usuario}' reseteada por {admin_usuario}", admin_usuario)
                print(Fore.GREEN + "Contraseña reseteada. El usuario deberá cambiarla al iniciar sesión.")
            else:
                print(Fore.RED + "La contraseña no cumple los requisitos de longitud.")
        elif opcion == '3':
            target_user_obj.is_active = not target_user_obj.is_active
            db_manager.update_user(target_user_obj)
            accion_log = "bloqueado" if not target_user_obj.is_active else "desbloqueado"
            registrar_movimiento("SISTEMA", "Estado Usuario", f"Acceso de '{target_user_obj.nombre_usuario}' {accion_log} por {admin_usuario}", admin_usuario)
            print(Fore.GREEN + f"Acceso {accion_log}.")
        elif opcion == '4':
            break
        else:
            print(Fore.RED + "Opción no válida.")
        pausar_pantalla()

@requiere_permiso("ver_historico")
def menu_ver_historico(usuario: str):
    while True:
        mostrar_menu(["Ver en Consola", "Exportar a Excel", "Volver"], titulo="Histórico de Movimientos")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        if opcion == '1': ver_historico_consola()
        elif opcion == '2': generar_excel_historico(usuario)
        elif opcion == '3': break
        else: print(Fore.RED + "Opción no válida.")

def ver_historico_consola():
    mostrar_encabezado("Histórico de Movimientos")
    historico = db_manager.get_all_historico()
    if not historico:
        print(Fore.YELLOW + "\nNo hay registros históricos.")
    else:
        for mov in historico:
            print(f"{Fore.CYAN}Fecha:{Style.RESET_ALL} {mov['fecha']} | {Fore.CYAN}Equipo:{Style.RESET_ALL} {mov['equipo_placa']} | {Fore.CYAN}Acción:{Style.RESET_ALL} {mov['accion']} | {Fore.CYAN}Usuario:{Style.RESET_ALL} {mov['usuario']}")
            print(f"  {Fore.CYAN}Detalles:{Style.RESET_ALL} {mov['detalles']}")
            print("-" * 60)
    pausar_pantalla()

@requiere_permiso("ver_inventario")
def ver_inventario_consola():
    mostrar_encabezado("Inventario Actual de Equipos")
    inventario = db_manager.get_all_equipos()
    if not inventario:
        print(Fore.YELLOW + "\nEl inventario está vacío.")
    else:
        print(f"{Fore.CYAN}{'PLACA':<12} {'TIPO':<15} {'MARCA':<15} {'MODELO':<20} {'ESTADO':<30} {'ASIGNADO A':<20}{Style.RESET_ALL}")
        print(Fore.CYAN + "="*112 + Style.RESET_ALL)
        for equipo in inventario:
            estado_color = Fore.WHITE
            if equipo['estado'] == "Disponible": estado_color = Fore.GREEN
            elif equipo['estado'] in ["Asignado", "En préstamo"]: estado_color = Fore.YELLOW
            elif equipo['estado'] == "En mantenimiento": estado_color = Fore.MAGENTA
            elif equipo['estado'] == "Pendiente Devolución a Proveedor": estado_color = Fore.LIGHTYELLOW_EX
            elif equipo['estado'] == "Devuelto a Proveedor": estado_color = Fore.LIGHTBLACK_EX
            elif equipo['estado'] == "Dado de baja": estado_color = Fore.RED
            asignado_a = equipo.get('asignado_a') or 'N/A'
            print(f"{equipo['placa']:<12} {equipo['tipo']:<15} {equipo['marca']:<15} {equipo['modelo']:<20} {estado_color}{equipo['estado']:<30}{Style.RESET_ALL} {asignado_a:<20}")
    pausar_pantalla()

@requiere_permiso("configurar_sistema")
def menu_configuracion_sistema(usuario: str):
    while True:
        mostrar_menu(["Gestionar Tipos de Equipo", "Gestionar Marcas", "Volver"], titulo="Configuración del Sistema")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        if opcion == '1':
            gestionar_parametros(usuario, 'tipo_equipo', 'Tipo de Equipo')
        elif opcion == '2':
            gestionar_parametros(usuario, 'marca_equipo', 'Marca')
        elif opcion == '3':
            break
        else:
            print(Fore.RED + "Opción no válida.")

def gestionar_parametros(usuario: str, tipo_parametro: str, nombre_amigable: str):
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        mostrar_encabezado(f"Gestionar {nombre_amigable}s")
        
        items = db_manager.get_parametros_por_tipo(tipo_parametro)
        if not items:
            print(Fore.YELLOW + f"No hay {nombre_amigable}s configurados.")
        else:
            for i, item in enumerate(items, 1):
                estado = Fore.GREEN + "[Activo]" if item['is_active'] else Fore.RED + "[Inactivo]"
                print(f"{i}. {item['valor']} {estado}{Style.RESET_ALL}")
        
        print("\n")
        mostrar_menu([f"Añadir nuevo {nombre_amigable}", f"Activar/Inactivar un {nombre_amigable}", "Volver"], titulo="Opciones")
        
        opcion = input(Fore.YELLOW + "Seleccione: " + Style.RESET_ALL).strip()

        if opcion == '1':
            nuevo_valor = input(Fore.YELLOW + f"Ingrese el nuevo {nombre_amigable}: " + Style.RESET_ALL).strip()
            if nuevo_valor:
                try:
                    db_manager.add_parametro(tipo_parametro, nuevo_valor)
                    registrar_movimiento("SISTEMA", "Configuración", f"Añadido {nombre_amigable}: '{nuevo_valor}'", usuario)
                    print(Fore.GREEN + f"{nombre_amigable} '{nuevo_valor}' añadido con éxito.")
                except sqlite3.IntegrityError:
                    print(Fore.RED + f"El {nombre_amigable} '{nuevo_valor}' ya existe.")
            else:
                print(Fore.RED + "El valor no puede estar vacío.")
            pausar_pantalla()
        elif opcion == '2':
            if not items:
                print(Fore.RED + f"No hay {nombre_amigable}s para gestionar.")
                pausar_pantalla()
                continue
            
            valor_a_gestionar = input(Fore.YELLOW + f"Ingrese el nombre exacto del {nombre_amigable} a activar/inactivar: " + Style.RESET_ALL).strip()
            
            item_encontrado = next((item for item in items if item['valor'] == valor_a_gestionar), None)

            if item_encontrado:
                es_activo_actualmente = item_encontrado['is_active']
                
                if es_activo_actualmente:
                    if db_manager.is_parametro_in_use(tipo_parametro, valor_a_gestionar):
                        print(Fore.RED + f"\n❌ No se puede inactivar '{valor_a_gestionar}'. Está siendo utilizado por al menos un equipo.")
                        pausar_pantalla()
                        continue 

                nuevo_estado = not es_activo_actualmente
                db_manager.update_parametro_status(tipo_parametro, valor_a_gestionar, nuevo_estado)
                accion = "activado" if nuevo_estado else "inactivado"
                registrar_movimiento("SISTEMA", "Configuración", f"{nombre_amigable} '{valor_a_gestionar}' {accion}", usuario)
                print(Fore.GREEN + f"\n✅ {nombre_amigable} '{valor_a_gestionar}' {accion} con éxito.")
            else:
                print(Fore.RED + f"El {nombre_amigable} '{valor_a_gestionar}' no fue encontrado.")
            pausar_pantalla()
        elif opcion == '3':
            break
        else:
            print(Fore.RED + "Opción no válida.")
            pausar_pantalla()

# --- BUCLE PRINCIPAL ---
def menu_principal():
    global USUARIO_ACTUAL
    inicializar_admin_si_no_existe()

    while USUARIO_ACTUAL is None:
        USUARIO_ACTUAL = login()
        if USUARIO_ACTUAL is None:
            if input(Fore.RED + "¿Salir del programa? (S/N): " + Style.RESET_ALL).strip().upper() == 'S':
                return

    user_data = db_manager.get_user_by_username(USUARIO_ACTUAL)
    rol_actual = user_data['rol']

    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        print(Fore.BLUE + "═" * 80)
        print(Back.BLUE + Fore.WHITE + f" MENÚ PRINCIPAL - Usuario: {USUARIO_ACTUAL} (Rol: {rol_actual}) ".center(80, ' ') + Style.RESET_ALL)
        print(Fore.BLUE + "═" * 80 + Style.RESET_ALL)
        
        opciones_principales = ["Gestión de Inventario", "Gestión de Acceso y Sistema", "Salir"]
        mostrar_menu(opciones_principales, titulo="Módulos del Sistema")
        
        opcion = input(Fore.YELLOW + "Seleccione un módulo: " + Style.RESET_ALL).strip()
        if opcion == '1':
            menu_gestion_inventario(USUARIO_ACTUAL)
        elif opcion == '2':
            menu_gestion_acceso_sistema(USUARIO_ACTUAL)
        elif opcion == '3':
            break
        else:
            print(Fore.RED + "\n❌ Opción no válida.")
            pausar_pantalla()
    
    print(Fore.GREEN + "\n¡Gracias por usar el Sistema de Gestión de Inventario!")

if __name__ == "__main__":
    try:
        menu_principal()
    except KeyboardInterrupt:
        print(Fore.RED + "\n\nPrograma interrumpido por el usuario.")
    except Exception as e:
        print(Fore.RED + f"\n\n❌ Un error inesperado ha ocurrido: {str(e)}")
    finally:
        db_manager.close()
        print(Fore.GREEN + "\nConexión a la base de datos cerrada.")
