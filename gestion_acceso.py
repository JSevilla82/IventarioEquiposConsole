# gestion_acceso.py
import os
import bcrypt
import getpass
import re
import time
from typing import Callable, Dict
import sqlite3
import webbrowser
import tempfile
from datetime import datetime
from functools import wraps

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from colorama import Fore, Back, Style

from database import db_manager, Usuario, registrar_movimiento_sistema
import ui

# --- CONTROL DE ACCESO BASADO EN ROLES (RBAC) ---
ROLES_PERMISOS = {
    "Administrador": {
        "registrar_equipo", "ver_inventario", "gestionar_equipo", "ver_historico",
        "generar_reporte", "gestionar_usuarios", "eliminar_equipo",
        "devolver_a_proveedor", "aprobar_devoluciones", "gestionar_pendientes",
        "configurar_sistema"
    },
    "Gestor": {
        "registrar_equipo", "ver_inventario", "gestionar_equipo", "ver_historico",
        "generar_reporte", "devolver_a_proveedor"
    },
    "Visualizador": {
        "ver_inventario", "ver_historico", "generar_reporte"
    }
}

def requiere_permiso(permiso: str) -> Callable:
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            if ui.USUARIO_ACTUAL is None:
                print(Fore.RED + "\n❌ Acceso denegado. No hay usuario logueado." + Style.RESET_ALL)
                return
            user_data = db_manager.get_user_by_username(ui.USUARIO_ACTUAL)
            if not user_data:
                print(Fore.RED + "\n❌ Acceso denegado. Usuario no encontrado." + Style.RESET_ALL)
                return
            rol_usuario = user_data['rol']
            if permiso in ROLES_PERMISOS.get(rol_usuario, {}):
                return func(*args, **kwargs)
            else:
                print(Fore.RED + f"\n❌ Permiso denegado. Su rol '{rol_usuario}' no tiene el permiso '{permiso}'." + Style.RESET_ALL)
                ui.pausar_pantalla()
                return
        return wrapper
    return decorator

# --- FUNCIONES DE AUTENTICACIÓN Y GESTIÓN DE USUARIOS ---
def hash_contrasena(contrasena: str) -> str:
    return bcrypt.hashpw(contrasena.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verificar_contrasena(contrasena: str, hash_almacenado: str) -> bool:
    return bcrypt.checkpw(contrasena.encode('utf-8'), hash_almacenado.encode('utf-8'))

def validar_contrasena(contrasena: str) -> bool:
    if len(contrasena) < 8:
        return False
    if not re.search(r'[A-Za-z]', contrasena):
        return False
    if not re.search(r'[0-9]', contrasena):
        return False
    return True

def login():
    error_message = ""
    nombre_usuario = ""
    intentos = 0
    
    while intentos < 3:
        ui.mostrar_encabezado("Inicio de Sesión")
        print("Bienvenido al Control de Inventario de Equipos (CIE).")
        print("Por favor, ingrese su usuario y contraseña para continuar.")
        print(Fore.WHITE + "─" * 80)

        if error_message:
            print(Fore.RED + f"\n{error_message}\n")
        
        nombre_usuario = ui.solicitar_input(Fore.YELLOW + "👤 Ingrese su usuario: ", default=nombre_usuario)
        contrasena = ui.solicitar_contrasena_con_asteriscos(Fore.YELLOW + "🔑 Ingrese su contraseña: ")

        if not nombre_usuario or not contrasena:
            error_message = "❌ El usuario y la contraseña no pueden estar vacíos."
            continue
        
        # --- Secuencia de carga ---
        loading_messages = [
            "Verificando credenciales...",
        ]
        total_delay = 2  # segundos
        delay_per_message = total_delay / len(loading_messages)

        print() # Espacio antes de los mensajes de carga
        for msg in loading_messages:
            print(Fore.CYAN + f"\r{msg.ljust(40)}", end="", flush=True)
            time.sleep(delay_per_message)
        
        print("\r" + " " * 40 + "\r", end="", flush=True) # Limpia la línea
        # --- Fin de la secuencia de carga ---

        user_data = db_manager.get_user_by_username(nombre_usuario)
        
        if user_data and verificar_contrasena(contrasena, user_data['contrasena_hash']):
            if not user_data['is_active']:
                error_message = "❌ Su cuenta de usuario está bloqueada. Contacte a un administrador."
                intentos += 1
                continue
            
            # --- Inicio de sesión exitoso ---
            if user_data.get('cambio_clave_requerido'):
                ui.mostrar_encabezado("Cambio de Contraseña Requerido")
                print(Fore.YELLOW + "⚠️ Su contraseña ha expirado y debe ser cambiada.")
                ui.pausar_pantalla()
                cambiar_contrasena_usuario(nombre_usuario, forzar_cambio=True)
            
            # Se retorna directamente para ir al menú principal
            return nombre_usuario
        else:
            error_message = "❌ Credenciales incorrectas."
            intentos += 1

    print(Fore.RED + "\n❌ Demasiados intentos fallidos.")
    return None


@requiere_permiso("gestionar_usuarios")
def registrar_usuario(usuario_actual: str):
    ui.mostrar_encabezado("Registrar Nuevo Usuario")
    print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar.")
    try:
        nombre_completo = input(Fore.YELLOW + "Nombre completo: " + Style.RESET_ALL).strip()
        nombre_usuario = input(Fore.YELLOW + "Nombre de usuario: " + Style.RESET_ALL).strip().lower()
        if not nombre_completo or not nombre_usuario:
            print(Fore.RED + "Nombre completo y nombre de usuario son obligatorios."); return
        if db_manager.get_user_by_username(nombre_usuario):
            print(Fore.RED + "Este nombre de usuario ya existe."); return
        
        contrasena = getpass.getpass(Fore.YELLOW + "Contraseña (mín. 8 caracteres): " + Style.RESET_ALL)
        if not validar_contrasena(contrasena):
            print(Fore.RED + "La contraseña es muy corta."); return

        roles_validos = ["Administrador", "Gestor", "Visualizador"]
        rol = input(Fore.YELLOW + f"Rol ({', '.join(roles_validos)}): " + Style.RESET_ALL).strip().capitalize()
        if rol not in roles_validos:
            print(Fore.RED + "Rol no válido."); return

        nuevo_usuario = Usuario(nombre_usuario, hash_contrasena(contrasena), rol, nombre_completo, True, True)
        db_manager.insert_user(nuevo_usuario)
        registrar_movimiento_sistema("Registro Usuario", f"Usuario '{nombre_usuario}' ({rol}) registrado por {usuario_actual}", usuario_actual)
        print(Fore.GREEN + f"\n✅ Usuario '{nombre_usuario}' registrado.")
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        ui.pausar_pantalla()

def cambiar_contrasena_usuario(nombre_usuario: str, forzar_cambio: bool = False):
    error_message = ""
    while True:
        ui.mostrar_encabezado(f"Cambiar Contraseña para {nombre_usuario}")
        print(Fore.CYAN + "💡 La contraseña debe tener al menos 8 caracteres, incluyendo letras y números.")
        print(Fore.WHITE + "─" * 80)

        if error_message:
            print(Fore.RED + f"\n{error_message}\n")
            error_message = ""
        
        user_data = db_manager.get_user_by_username(nombre_usuario)
        if not user_data:
            print(Fore.RED + "Usuario no encontrado.")
            break

        try:
            if not forzar_cambio:
                old_password = ui.solicitar_contrasena_con_asteriscos(Fore.YELLOW + "🔑 Ingrese su contraseña actual: ")
                if not old_password:
                    error_message = "❌ La contraseña actual no puede estar vacía."
                    continue
                if not verificar_contrasena(old_password, user_data['contrasena_hash']):
                    error_message = "❌ Contraseña actual incorrecta."
                    continue
            
            new_password = ui.solicitar_contrasena_con_asteriscos(Fore.YELLOW + "🔑 Ingrese su nueva contraseña: ")
            if not validar_contrasena(new_password):
                error_message = "❌ La nueva contraseña es muy corta o no contiene letras y números."
                continue
                
            confirm_password = ui.solicitar_contrasena_con_asteriscos(Fore.YELLOW + "🔑 Confirme la nueva contraseña: ")
            if new_password != confirm_password:
                error_message = "❌ Las contraseñas no coinciden."
                continue

            user_obj = Usuario(**user_data)
            user_obj.contrasena_hash = hash_contrasena(new_password)
            user_obj.cambio_clave_requerido = False
            db_manager.update_user(user_obj)
            registrar_movimiento_sistema("Cambio Contraseña", f"Contraseña cambiada para '{nombre_usuario}'", nombre_usuario)
            print(Fore.GREEN + "\n✅ Contraseña cambiada exitosamente.")
            ui.pausar_pantalla()
            break

        except KeyboardInterrupt:
            print(Fore.CYAN + "\n🚫 Operación cancelada.")
            ui.pausar_pantalla()
            break

def inicializar_admin_si_no_existe():
    if not db_manager.get_user_by_username("admin"):
        print(Fore.YELLOW + "\nCreando usuario administrador inicial 'admin'...")
        admin_pass_hash = hash_contrasena("adminpass")
        admin_user = Usuario("admin", admin_pass_hash, "Administrador", "Administrador Principal", True, True)
        db_manager.insert_user(admin_user)
        print(Fore.GREEN + "✅ Usuario 'admin' creado con contraseña 'adminpass'. Por favor, cámbiela.")
        ui.pausar_pantalla()

@requiere_permiso("gestionar_usuarios")
def menu_usuarios(usuario_actual: str):
    while True:
        # La llamada a mostrar_encabezado se encarga de limpiar la pantalla.
        ui.mostrar_encabezado("Gestión de Usuarios")
        usuarios = db_manager.get_all_users()
        
        print(f"{Fore.CYAN}{'USUARIO':<20} {'NOMBRE COMPLETO':<30} {'ESTADO'}{Style.RESET_ALL}")
        print(Fore.CYAN + "-" * 65 + Style.RESET_ALL)
        if not usuarios:
            print(Fore.YELLOW + "No hay usuarios registrados.")
        else:
            for user in usuarios:
                estado = Fore.GREEN + "Activo" if user['is_active'] else Fore.RED + "Bloqueado"
                nombre_completo = user.get('nombre_completo') or 'N/A'
                print(f"{user['nombre_usuario']:<20} {nombre_completo:<30} {estado}{Style.RESET_ALL}")
        print(Fore.CYAN + "-" * 65 + Style.RESET_ALL)
        
        ui.mostrar_menu(["Registrar nuevo usuario", "Gestionar un usuario existente", "Volver"], titulo="Opciones de Gestión de Usuarios")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()

        if opcion == '1':
            registrar_usuario(usuario_actual)
        elif opcion == '2':
            if not usuarios:
                print(Fore.YELLOW + "No hay usuarios para gestionar.")
                ui.pausar_pantalla()
                continue
            target_username = input(Fore.YELLOW + "Ingrese el nombre de usuario a gestionar: " + Style.RESET_ALL).strip().lower()
            target_user_data = db_manager.get_user_by_username(target_username)
            if target_user_data:
                gestionar_usuario_especifico(usuario_actual, target_user_data)
            else:
                print(Fore.RED + "Usuario no encontrado.")
                ui.pausar_pantalla()
        elif opcion == '3':
            break
        else:
            print(Fore.RED + "Opción no válida.")
            ui.pausar_pantalla()

def gestionar_usuario_especifico(admin_usuario: str, target_user_data: Dict):
    target_user_obj = Usuario(**db_manager.get_user_by_username(target_user_data['nombre_usuario']))
    
    while True:
        estado = Fore.GREEN + "Activo" if target_user_obj.is_active else Fore.RED + "Bloqueado"
        ui.mostrar_encabezado(f"Gestionando a {target_user_obj.nombre_usuario}", color=Fore.GREEN)
        
        opcion_bloqueo = (Fore.RED + "Bloquear Acceso") if target_user_obj.is_active else (Fore.GREEN + "Desbloquear Acceso")
        
        opciones_menu = [
            "Modificar nombre completo",
            "Resetear contraseña",
            opcion_bloqueo,
            "Volver"
        ]
        
        titulo_menu = f"OPCIONES - {target_user_obj.nombre_completo or 'N/A'} / {target_user_obj.nombre_usuario}"
        ui.mostrar_menu(opciones_menu, titulo=titulo_menu)
        opcion = input(Fore.YELLOW + "Seleccione: " + Style.RESET_ALL).strip()

        if opcion == '1':
            nuevo_nombre = input(Fore.YELLOW + f"Nuevo nombre completo ({target_user_obj.nombre_completo}): " + Style.RESET_ALL).strip()
            if nuevo_nombre:
                target_user_obj.nombre_completo = nuevo_nombre
                db_manager.update_user(target_user_obj)
                registrar_movimiento_sistema("Modificación Usuario", f"Nombre de '{target_user_obj.nombre_usuario}' cambiado a '{nuevo_nombre}' por {admin_usuario}", admin_usuario)
                print(Fore.GREEN + "Nombre actualizado.")
        elif opcion == '2':
            nueva_pass = getpass.getpass(Fore.YELLOW + f"Ingrese la nueva contraseña para {target_user_obj.nombre_usuario}: " + Style.RESET_ALL)
            if validar_contrasena(nueva_pass):
                target_user_obj.contrasena_hash = hash_contrasena(nueva_pass)
                target_user_obj.cambio_clave_requerido = True
                db_manager.update_user(target_user_obj)
                registrar_movimiento_sistema("Reset Contraseña", f"Contraseña de '{target_user_obj.nombre_usuario}' reseteada por {admin_usuario}", admin_usuario)
                print(Fore.GREEN + "Contraseña reseteada. El usuario deberá cambiarla al iniciar sesión.")
            else:
                print(Fore.RED + "La contraseña no cumple los requisitos de longitud.")
        elif opcion == '3':
            target_user_obj.is_active = not target_user_obj.is_active
            db_manager.update_user(target_user_obj)
            accion_log = "bloqueado" if not target_user_obj.is_active else "desbloqueado"
            registrar_movimiento_sistema("Estado Usuario", f"Acceso de '{target_user_obj.nombre_usuario}' {accion_log} por {admin_usuario}", admin_usuario)
            print(Fore.GREEN + f"Acceso {accion_log}.")
        elif opcion == '4':
            break
        else:
            print(Fore.RED + "Opción no válida.")
        ui.pausar_pantalla()

@requiere_permiso("ver_historico")
def menu_ver_log_sistema(usuario: str):
    while True:
        ui.mostrar_menu(["Ver Log en Excel", "Volver"], titulo="Log de Actividad del Sistema")
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        if opcion == '1':
            generar_excel_log_sistema(usuario)
        elif opcion == '2':
            break
        else:
            print(Fore.RED + "Opción no válida.")

@requiere_permiso("ver_historico")
def generar_excel_log_sistema(usuario: str):
    try:
        log_sistema = db_manager.get_all_log_sistema()

        if not log_sistema:
            print(Fore.YELLOW + "\nNo hay actividad del sistema para exportar.")
            ui.pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Log del Sistema"

        header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = ["FECHA", "ACCIÓN", "USUARIO", "DETALLES"]
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 80

        for row_num, mov in enumerate(log_sistema, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('detalles', '')).border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        print(Fore.GREEN + f"\n✅ Abriendo el log de actividad del sistema en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el log del sistema: {str(e)}" + Style.RESET_ALL)
    finally:
        ui.pausar_pantalla()

# --- NUEVA FUNCIÓN ---
@requiere_permiso("configurar_sistema")
def gestionar_proveedores(usuario: str):
    """Permite gestionar los proveedores y las iniciales de sus placas."""
    while True:
        ui.mostrar_encabezado("Gestionar Proveedores")
        
        proveedores = db_manager.get_all_proveedores()
        if not proveedores:
            print(Fore.YELLOW + "No hay proveedores configurados.")
        else:
            for i, prov in enumerate(proveedores, 1):
                # --- MODIFICACIÓN: Mostrar estado Activo/Inactivo ---
                estado = Fore.GREEN + "[Activo]" if prov.get('is_active', 1) else Fore.RED + "[Inactivo]"
                en_uso = Fore.RED + " (En uso)" if db_manager.is_proveedor_in_use(prov['nombre']) else ""
                inicial_placa = f" (Inicial Placa: {prov['placa_inicial']})" if prov['placa_inicial'] else ""
                print(f"{i}. {prov['nombre']}{inicial_placa} {estado}{en_uso}{Style.RESET_ALL}")
        
        print("\n")
        # --- MODIFICACIÓN: Añadir nueva opción al menú ---
        opciones_menu = [
            "Añadir nuevo proveedor",
            "Activar/Inactivar un proveedor",
            "Eliminar un proveedor",
            "Volver"
        ]
        ui.mostrar_menu(opciones_menu, titulo="Opciones")
        
        opcion = input(Fore.YELLOW + "Seleccione: " + Style.RESET_ALL).strip()

        if opcion == '1':
            try:
                # --- Formulario interactivo ---
                ui.mostrar_encabezado("Registrar Nuevo Proveedor")
                print(Fore.CYAN + "💡 Ingrese los datos del nuevo proveedor.")
                
                nombre_proveedor = input(Fore.YELLOW + "Paso 1: Ingrese el nombre del nuevo proveedor: " + Style.RESET_ALL).strip().title()
                if not nombre_proveedor:
                    print(Fore.RED + "El nombre no puede estar vacío."); continue
                
                print(Fore.CYAN + "\n💡 Si este proveedor identifica sus equipos con un identificador que siempre empieza con las mismas iniciales, configúrelas a continuación.")
                print(Fore.CYAN + "   De lo contrario, deje este campo en blanco.")
                print(Fore.CYAN + "   Ejemplo: ML-12345 (Todas las placas llevan 'ML-' al inicio)")
                
                placa_inicial = input(Fore.YELLOW + "Paso 2: Ingrese las iniciales de la placa (opcional): " + Style.RESET_ALL).strip().upper()

                print("\n" + Fore.CYAN + "--- Resumen para Confirmación ---")
                print(f"Nombre del Proveedor: {nombre_proveedor}")
                print(f"Iniciales de Placa:   {placa_inicial or 'Ninguna'}")
                print("---------------------------------" + Style.RESET_ALL)

                confirmacion_final = input(Fore.YELLOW + "¿Desea registrar este nuevo proveedor? (S/N): " + Style.RESET_ALL).strip().upper()
                if confirmacion_final == 'S':
                    db_manager.insert_proveedor(nombre_proveedor, placa_inicial)
                    registrar_movimiento_sistema("Configuración", f"Añadido Proveedor: '{nombre_proveedor}'", usuario)
                    print(Fore.GREEN + f"\n✅ Proveedor '{nombre_proveedor}' añadido con éxito.")
                else:
                    print(Fore.YELLOW + "\nOperación cancelada.")
            except sqlite3.IntegrityError:
                print(Fore.RED + f"\n❌ El proveedor '{nombre_proveedor}' ya existe.")
            except KeyboardInterrupt:
                print(Fore.CYAN + "\n🚫 Operación cancelada.")
            finally:
                ui.pausar_pantalla()

        elif opcion == '2': # --- NUEVA LÓGICA ---
            if not proveedores:
                print(Fore.RED + "No hay proveedores para gestionar."); continue
            
            valor_a_gestionar = input(Fore.YELLOW + "Ingrese el nombre exacto del proveedor a activar/inactivar: " + Style.RESET_ALL).strip().title()
            item_encontrado = db_manager.get_proveedor_by_name(valor_a_gestionar)

            if item_encontrado:
                es_activo_actualmente = item_encontrado.get('is_active', 1)
                
                if es_activo_actualmente and db_manager.is_proveedor_in_use(valor_a_gestionar):
                    print(Fore.RED + f"\n❌ No se puede inactivar '{valor_a_gestionar}'. Está siendo utilizado por al menos un equipo.")
                    ui.pausar_pantalla()
                    continue 

                nuevo_estado = not es_activo_actualmente
                db_manager.update_proveedor_status(valor_a_gestionar, nuevo_estado)
                accion_log = "activado" if nuevo_estado else "inactivado"
                registrar_movimiento_sistema("Configuración", f"Proveedor '{valor_a_gestionar}' {accion_log}", usuario)
                print(Fore.GREEN + f"\n✅ Proveedor '{valor_a_gestionar}' {accion_log} con éxito.")
            else:
                print(Fore.RED + f"El proveedor '{valor_a_gestionar}' no fue encontrado.")
            ui.pausar_pantalla()

        elif opcion == '3':
            if not proveedores:
                print(Fore.RED + "No hay proveedores para eliminar."); continue

            valor_a_eliminar = input(Fore.YELLOW + "Ingrese el nombre exacto del proveedor a ELIMINAR: " + Style.RESET_ALL).strip().title()
            
            if not db_manager.get_proveedor_by_name(valor_a_eliminar):
                print(Fore.RED + f"El proveedor '{valor_a_eliminar}' no fue encontrado."); continue

            if db_manager.is_proveedor_in_use(valor_a_eliminar):
                print(Fore.RED + f"\n❌ No se puede eliminar '{valor_a_eliminar}'. Está siendo utilizado por al menos un equipo.")
                ui.pausar_pantalla()
                continue
            
            confirmacion = input(Fore.RED + f"⚠️ ¿Seguro de eliminar el proveedor '{valor_a_eliminar}'? Esta acción es irreversible. (Escriba 'SI'): " + Style.RESET_ALL).strip().upper()
            if confirmacion == "SI":
                db_manager.delete_proveedor(valor_a_eliminar)
                registrar_movimiento_sistema("Configuración", f"Eliminado Proveedor: '{valor_a_eliminar}'", usuario)
                print(Fore.GREEN + f"\n✅ Proveedor '{valor_a_eliminar}' eliminado con éxito.")
            else:
                print(Fore.YELLOW + "\nOperación de eliminación cancelada.")
            ui.pausar_pantalla()

        elif opcion == '4':
            break
        else:
            print(Fore.RED + "Opción no válida.")
            ui.pausar_pantalla()

# --- FUNCIÓN MODIFICADA ---
def gestionar_parametros(usuario: str, tipo_parametro: str, nombre_amigable: str):
    while True:
        ui.mostrar_encabezado(f"Gestionar {nombre_amigable}s")
        
        items = db_manager.get_parametros_por_tipo(tipo_parametro)
        if not items:
            print(Fore.YELLOW + f"No hay {nombre_amigable}s configurados.")
        else:
            for i, item in enumerate(items, 1):
                estado = Fore.GREEN + "[Activo]" if item['is_active'] else Fore.RED + "[Inactivo]"
                en_uso = Fore.RED + " (En uso)" if db_manager.is_parametro_in_use(tipo_parametro, item['valor']) else ""
                print(f"{i}. {item['valor']} {estado}{en_uso}{Style.RESET_ALL}")
        
        print("\n")
        opciones_menu = [
            f"Añadir nuevo {nombre_amigable}",
            f"Activar/Inactivar un {nombre_amigable}",
            f"Eliminar un {nombre_amigable}",
            "Volver"
        ]
        ui.mostrar_menu(opciones_menu, titulo="Opciones")
        
        opcion = input(Fore.YELLOW + "Seleccione: " + Style.RESET_ALL).strip()
        
        opciones_validas = {"1": "add", "2": "toggle", "3": "delete", "4": "back"}
        accion = opciones_validas.get(opcion)

        if accion == "add":
            try:
                while True:
                    nuevo_valor = input(Fore.YELLOW + f"Paso 1: Ingrese el nuevo {nombre_amigable}: " + Style.RESET_ALL).strip()
                    if not nuevo_valor:
                        print(Fore.RED + "El valor no puede estar vacío.")
                        continue
                    
                    # --- APLICAR FORMATO ---
                    if tipo_parametro == 'dominio_correo':
                        nuevo_valor = nuevo_valor.lower()
                        if '@' in nuevo_valor or '.' not in nuevo_valor:
                            print(Fore.RED + "Formato de dominio inválido. Ingrese solo el dominio (ej: gmail.com).")
                            continue
                    else:
                        nuevo_valor = nuevo_valor.title()
                    
                    confirmacion_valor = input(Fore.YELLOW + f"Paso 2: Confirme el nuevo {nombre_amigable}: " + Style.RESET_ALL).strip()
                    
                    # --- APLICAR FORMATO A LA CONFIRMACIÓN ---
                    if tipo_parametro == 'dominio_correo':
                        confirmacion_valor = confirmacion_valor.lower()
                    else:
                        confirmacion_valor = confirmacion_valor.title()

                    if nuevo_valor == confirmacion_valor:
                        break
                    else:
                        print(Fore.RED + "Los valores no coinciden. Por favor, intente de nuevo.")
                
                print("\n" + Fore.CYAN + "--- Resumen para Confirmación ---")
                print(f"Tipo de Parámetro: {nombre_amigable}")
                print(f"Valor a registrar:   {nuevo_valor}")
                print("---------------------------------" + Style.RESET_ALL)
                
                confirmacion_final = input(Fore.YELLOW + f"¿Desea registrar este nuevo {nombre_amigable}? (S/N): " + Style.RESET_ALL).strip().upper()

                if confirmacion_final == 'S':
                    try:
                        db_manager.add_parametro(tipo_parametro, nuevo_valor)
                        registrar_movimiento_sistema("Configuración", f"Añadido {nombre_amigable}: '{nuevo_valor}'", usuario)
                        print(Fore.GREEN + f"\n✅ {nombre_amigable} '{nuevo_valor}' añadido con éxito.")
                    except sqlite3.IntegrityError:
                        print(Fore.RED + f"\n❌ El {nombre_amigable} '{nuevo_valor}' ya existe.")
                else:
                    print(Fore.YELLOW + "\nOperación cancelada.")
            except KeyboardInterrupt:
                print(Fore.CYAN + "\n🚫 Operación cancelada.")
            finally:
                ui.pausar_pantalla()

        elif accion == "toggle":
            if not items:
                print(Fore.RED + f"No hay {nombre_amigable}s para gestionar.")
                ui.pausar_pantalla()
                continue
            
            valor_a_gestionar = input(Fore.YELLOW + f"Ingrese el nombre exacto del {nombre_amigable} a activar/inactivar: " + Style.RESET_ALL).strip()
            item_encontrado = next((item for item in items if item['valor'] == valor_a_gestionar), None)

            if item_encontrado:
                es_activo_actualmente = item_encontrado['is_active']
                
                if es_activo_actualmente and db_manager.is_parametro_in_use(tipo_parametro, valor_a_gestionar):
                    print(Fore.RED + f"\n❌ No se puede inactivar '{valor_a_gestionar}'. Está siendo utilizado por al menos un equipo.")
                    ui.pausar_pantalla()
                    continue 

                nuevo_estado = not es_activo_actualmente
                db_manager.update_parametro_status(tipo_parametro, valor_a_gestionar, nuevo_estado)
                accion_log = "activado" if nuevo_estado else "inactivado"
                registrar_movimiento_sistema("Configuración", f"{nombre_amigable} '{valor_a_gestionar}' {accion_log}", usuario)
                print(Fore.GREEN + f"\n✅ {nombre_amigable} '{valor_a_gestionar}' {accion_log} con éxito.")
            else:
                print(Fore.RED + f"El {nombre_amigable} '{valor_a_gestionar}' no fue encontrado.")
            ui.pausar_pantalla()

        elif accion == "delete":
            if not items:
                print(Fore.RED + f"No hay {nombre_amigable}s para eliminar.")
                ui.pausar_pantalla()
                continue
            
            valor_a_eliminar = input(Fore.YELLOW + f"Ingrese el nombre exacto del {nombre_amigable} a ELIMINAR: " + Style.RESET_ALL).strip()
            item_encontrado = next((item for item in items if item['valor'] == valor_a_eliminar), None)

            if not item_encontrado:
                print(Fore.RED + f"El {nombre_amigable} '{valor_a_eliminar}' no fue encontrado.")
                ui.pausar_pantalla()
                continue

            if db_manager.is_parametro_in_use(tipo_parametro, valor_a_eliminar):
                print(Fore.RED + f"\n❌ No se puede eliminar '{valor_a_eliminar}'. Está siendo utilizado por al menos un equipo.")
                ui.pausar_pantalla()
                continue
            
            confirmacion = input(Fore.RED + f"⚠️ ¿Seguro de eliminar el parámetro '{valor_a_eliminar}'? Esta acción es irreversible. (Escriba 'SI'): " + Style.RESET_ALL).strip().upper()
            if confirmacion == "SI":
                db_manager.delete_parametro(tipo_parametro, valor_a_eliminar)
                registrar_movimiento_sistema("Configuración", f"Eliminado {nombre_amigable}: '{valor_a_eliminar}'", usuario)
                print(Fore.GREEN + f"\n✅ Parámetro '{valor_a_eliminar}' eliminado con éxito.")
            else:
                print(Fore.YELLOW + "\nOperación de eliminación cancelada.")
            ui.pausar_pantalla()
        
        elif accion == "back":
            break
        else:
            print(Fore.RED + "Opción no válida. Por favor, intente de nuevo.")
            ui.pausar_pantalla()