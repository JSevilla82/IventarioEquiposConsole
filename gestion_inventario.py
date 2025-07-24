# gestion_inventario.py
import os
import re
import webbrowser
from datetime import datetime
from typing import Optional, List

import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from colorama import Fore, Style

from database import db_manager, Equipo, registrar_movimiento_inventario, registrar_movimiento_sistema
from ui import mostrar_encabezado, mostrar_menu, pausar_pantalla, confirmar_con_placa
from gestion_acceso import requiere_permiso

# --- FUNCIONES DE UTILIDAD Y VALIDACIÓN ---
def validar_placa_unica(placa: str) -> bool:
    return db_manager.get_equipo_by_placa(placa) is None

def validar_email(email: str) -> bool:
    return re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email) is not None

def validar_placa_formato(placa: str) -> bool:
    return len(placa) >= 4 and placa.isalnum()

def validar_formato_fecha(fecha_str: str) -> Optional[datetime]:
    try:
        return datetime.strptime(fecha_str, "%d/%m/%Y")
    except ValueError:
        return None

def validar_campo_general(texto: str) -> bool:
    if not texto: return False
    return re.match(r'^[A-Za-z0-9\s\-_.,()]+$', texto) is not None

def validar_serial(serial: str) -> bool:
    if not serial: return False
    return serial.isalnum()
    
# --- NUEVA FUNCIÓN PARA VER MOVIMIENTOS ---
def menu_ver_ultimos_movimientos(usuario: str):
    """Muestra una tabla con los últimos 20 movimientos de inventario del usuario."""
    os.system('cls' if os.name == 'nt' else 'clear')
    movimientos = db_manager.get_last_movimientos_by_user(usuario, limit=20)
    
    mostrar_encabezado("Tus Últimos 20 Movimientos", color=Fore.BLUE)

    if not movimientos:
        print(Fore.YELLOW + "No has registrado movimientos recientemente.".center(80) + Style.RESET_ALL)
    else:
        print(f"{Fore.CYAN}{'FECHA':<17} {'PLACA':<12} {'MARCA':<15} {'ACCIÓN':<30}{Style.RESET_ALL}")
        print(Fore.CYAN + "-" * 74 + Style.RESET_ALL)
        
        for mov in movimientos:
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            accion = mov.get('accion', 'N/A')
            if len(accion) > 28:
                accion = accion[:27] + "..."

            placa = mov.get('equipo_placa', 'N/A')
            marca = mov.get('marca', 'N/A') or 'N/A'

            print(f"{fecha_formateada:<17} {placa:<12} {marca:<15} {accion:<30}")
    
    pausar_pantalla()

# --- MENÚ DE REPORTES ---
@requiere_permiso("ver_inventario")
def menu_ver_inventario_excel(usuario: str):
    """Muestra el menú para generar los reportes de inventario en Excel."""
    while True:
        mostrar_menu([
            "Reporte de Inventario Actual (Equipos Activos)",
            "Reporte de Equipos Devueltos a Proveedor",
            "Reporte Histórico Completo de Equipos (Log)",
            "Volver"
        ], titulo="Ver Inventario en Excel")
        
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()

        if opcion == '1':
            generar_excel_inventario(usuario)
        elif opcion == '2':
            generar_excel_devueltos_proveedor(usuario)
        elif opcion == '3':
            generar_excel_historico(usuario)
        elif opcion == '4':
            break
        else:
            print(Fore.RED + "Opción no válida.")

# --- FUNCIONES DE REPORTES (Excel) ---
@requiere_permiso("generar_reporte")
def generar_excel_inventario(usuario: str) -> None:
    """Genera un reporte Excel con los equipos activos (no incluye los devueltos al proveedor)."""
    try:
        inventario = db_manager.get_equipos_activos()
        if not inventario:
            print(Fore.YELLOW + "\nNo hay equipos activos para generar un reporte.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario de Equipos"

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = [
            "FECHA REGISTRO", "PLACA", "TIPO", "MARCA", "MODELO", "SERIAL", "ESTADO", 
            "FECHA ÚLTIMO CAMBIO", "USUARIO ÚLTIMO CAMBIO", "ASIGNADO A", "EMAIL", "ÚLTIMA OBSERVACIÓN"
        ]
        
        column_widths = {'A': 25, 'B': 15, 'C': 25, 'D': 25, 'E': 25, 'F': 30, 'G': 30, 'H': 25, 'I': 25, 'J': 30, 'K': 30, 'L': 80}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border

        colores_estado = {
            "Disponible": "C6EFCE", "Asignado": "FFEB9C", "En préstamo": "DDEBF7",
            "En mantenimiento": "FCE4D6", "Pendiente Devolución a Proveedor": "FFFFCC"
        }

        for row_num, equipo in enumerate(inventario, 2):
            ultimo_movimiento = db_manager.get_last_movimiento_by_placa(equipo['placa'])
            
            fecha_ult_cambio = "N/A"
            usuario_ult_cambio = "N/A"
            ultima_observacion = equipo.get('observaciones', 'N/A')

            if ultimo_movimiento:
                fecha_obj = datetime.strptime(ultimo_movimiento['fecha'], "%Y-%m-%d %H:%M:%S")
                fecha_ult_cambio = fecha_obj.strftime("%d/%m/%Y %H:%M")
                usuario_ult_cambio = ultimo_movimiento.get('usuario', 'N/A')
                ultima_observacion = ultimo_movimiento.get('detalles', ultima_observacion)

            data_row = [
                equipo.get('fecha_registro', 'N/A'), equipo.get('placa', 'N/A'), equipo.get('tipo', 'N/A'),
                equipo.get('marca', 'N/A'), equipo.get('modelo', 'N/A'), equipo.get('serial', 'N/A'),
                equipo.get('estado', 'N/A'), fecha_ult_cambio, usuario_ult_cambio,
                equipo.get('asignado_a', ''), equipo.get('email_asignado', ''), ultima_observacion
            ]
            
            for col_num, cell_value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
            
            estado_celda = ws.cell(row=row_num, column=7)
            color_hex = colores_estado.get(equipo.get('estado'))
            if color_hex:
                estado_celda.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
        ws.freeze_panes = "A2"
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Inventario Activo", f"Generado reporte con {len(inventario)} equipos", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte de inventario activo en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte Excel: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

@requiere_permiso("generar_reporte")
def generar_excel_devueltos_proveedor(usuario: str) -> None:
    try:
        inventario_devuelto = db_manager.get_equipos_devueltos()
        if not inventario_devuelto:
            print(Fore.YELLOW + "\nNo hay equipos devueltos al proveedor para reportar.")
            pausar_pantalla()
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipos Devueltos"
        
        header_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = [
            "PLACA", "TIPO", "MARCA", "MODELO", "SERIAL", 
            "FECHA DEVOLUCIÓN", "MOTIVO DEVOLUCIÓN", "ÚLTIMA OBSERVACIÓN"
        ]
        
        column_widths = {'A': 15, 'B': 25, 'C': 25, 'D': 25, 'E': 30, 'F': 25, 'G': 25, 'H': 80}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border

        for row_num, equipo in enumerate(inventario_devuelto, 2):
            data_row = [
                equipo.get('placa', 'N/A'), equipo.get('tipo', 'N/A'), equipo.get('marca', 'N/A'),
                equipo.get('modelo', 'N/A'), equipo.get('serial', 'N/A'),
                equipo.get('fecha_devolucion_proveedor', 'N/A'), equipo.get('motivo_devolucion', 'N/A'),
                equipo.get('observaciones', 'N/A')
            ]
            for col_num, cell_value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name
        
        registrar_movimiento_sistema("Reporte Equipos Devueltos", f"Generado reporte con {len(inventario_devuelto)} equipos devueltos", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte de equipos devueltos en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte de equipos devueltos: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

@requiere_permiso("ver_historico")
def generar_excel_historico(usuario: str):
    try:
        log_equipos = db_manager.get_all_log_inventario()

        if not log_equipos:
            print(Fore.YELLOW + "\nNo hay movimientos de equipos para exportar.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico de Movimientos"

        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = ["FECHA", "PLACA EQUIPO", "ACCIÓN", "USUARIO", "DETALLES"]
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 80

        for row_num, mov in enumerate(log_equipos, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('equipo_placa', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=5, value=mov.get('detalles', '')).border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Histórico Equipos", "Generado reporte de histórico de equipos", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el reporte histórico de equipos en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el reporte de histórico: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

def generar_excel_historico_equipo(usuario: str, equipo: Equipo):
    """Genera un reporte Excel con el historial de un solo equipo."""
    try:
        log_equipo = db_manager.get_log_by_placa(equipo.placa)

        if not log_equipo:
            print(Fore.YELLOW + f"\nNo hay historial para el equipo {equipo.placa}.")
            pausar_pantalla()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = f"Historial {equipo.placa}"

        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        encabezados = ["FECHA", "ACCIÓN", "USUARIO", "DETALLES"]
        
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center')
            celda.border = border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 80

        for row_num, mov in enumerate(log_equipo, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('detalles', '')).border = border
        
        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Histórico Individual", f"Generado reporte para placa {equipo.placa}", usuario)
        print(Fore.GREEN + f"\n✅ Abriendo el historial del equipo {equipo.placa} en Excel..." + Style.RESET_ALL)
        webbrowser.open(ruta_temporal)

    except Exception as e:
        print(Fore.RED + f"\n❌ Error al generar el historial del equipo: {str(e)}" + Style.RESET_ALL)
    finally:
        pausar_pantalla()

# --- FUNCIONES PRINCIPALES DE INVENTARIO ---
def seleccionar_parametro(tipo_parametro: Optional[str], nombre_amigable: str, lista_opciones: Optional[List[str]] = None) -> Optional[str]:
    parametros = lista_opciones if lista_opciones is not None else [p['valor'] for p in db_manager.get_parametros_por_tipo(tipo_parametro, solo_activos=True)]
    
    while True:
        print(Fore.GREEN + f"Seleccione un {nombre_amigable}:")
        for i, param in enumerate(parametros, 1):
            print(f"{i}. {param}")
        print() 

        seleccion = input(Fore.YELLOW + "Opción: " + Style.RESET_ALL).strip()
        try:
            idx = int(seleccion) - 1
            if 0 <= idx < len(parametros):
                return parametros[idx]
            else:
                print(Fore.RED + "Selección fuera de rango.")
        except ValueError:
            print(Fore.RED + "Por favor, ingrese un número.")

@requiere_permiso("registrar_equipo")
def registrar_equipo(usuario: str):
    mostrar_encabezado("Registro de Nuevo Equipo", color=Fore.BLUE)
    
    tipos_existentes = db_manager.get_parametros_por_tipo('tipo_equipo', solo_activos=True)
    marcas_existentes = db_manager.get_parametros_por_tipo('marca_equipo', solo_activos=True)

    if not tipos_existentes or not marcas_existentes:
        print(Fore.RED + "❌ No se puede registrar un nuevo equipo.")
        if not tipos_existentes: print(Fore.YELLOW + "   - No hay 'Tipos de Equipo' activos configurados.")
        if not marcas_existentes: print(Fore.YELLOW + "   - No hay 'Marcas' activas configuradas.")
        print(Fore.CYAN + "Por favor, pida a un Administrador que configure estos parámetros.")
        pausar_pantalla()
        return

    try:
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar." + Style.RESET_ALL)
        while True:
            placa = input(Fore.YELLOW + "Placa del equipo: " + Style.RESET_ALL).strip().upper()
            if not validar_placa_formato(placa):
                print(Fore.RED + "⚠️ Formato de placa inválido (mín. 4 caracteres alfanuméricos).")
                continue

            equipo_existente = db_manager.get_equipo_by_placa(placa)
            if equipo_existente:
                if equipo_existente['estado'] == "Devuelto a Proveedor":
                    print(Fore.YELLOW + f"\n⚠️ Este equipo (Placa: {placa}) ya existe y fue devuelto al proveedor.")
                    print(Fore.CYAN + "--- Información del Equipo Existente ---")
                    print(f"  Tipo: {equipo_existente['tipo']}, Marca: {equipo_existente['marca']}, Modelo: {equipo_existente['modelo']}")
                    print("---------------------------------------" + Style.RESET_ALL)
                    
                    confirmacion_reactivar = input(Fore.YELLOW + "¿Desea reactivar este equipo? (S/N): " + Style.RESET_ALL).strip().upper()
                    if confirmacion_reactivar == 'S':
                        if not confirmar_con_placa(placa):
                            print(Fore.YELLOW + "Reactivación cancelada.")
                            continue

                        equipo_reactivado = Equipo(**equipo_existente)
                        equipo_reactivado.estado = "Disponible"
                        equipo_reactivado.estado_anterior = "Devuelto a Proveedor"
                        equipo_reactivado.asignado_a = None
                        equipo_reactivado.email_asignado = None
                        equipo_reactivado.fecha_devolucion_proveedor = None
                        equipo_reactivado.motivo_devolucion = None
                        equipo_reactivado.fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        
                        db_manager.update_equipo(equipo_reactivado)
                        registrar_movimiento_inventario(placa, "Reactivación", "Equipo reactivado en el inventario tras devolución a proveedor.", usuario)
                        print(Fore.GREEN + f"\n✅ ¡Equipo {placa} reactivado y disponible en el inventario!")
                        pausar_pantalla()
                        return
                    else:
                        print(Fore.YELLOW + "Reactivación cancelada.")
                        continue
                else:
                    print(Fore.RED + "⚠️ Esta placa ya está registrada y activa en el sistema.")
            else:
                break
        
        tipo = seleccionar_parametro('tipo_equipo', 'Tipo de Equipo')
        marca = seleccionar_parametro('marca_equipo', 'Marca')

        while True:
            modelo = input(Fore.YELLOW + "Modelo: " + Style.RESET_ALL).strip()
            if validar_campo_general(modelo):
                break
            print(Fore.RED + "Modelo inválido. Solo se permiten letras, números, espacios y (- _ . ,).")
        
        while True:
            serial = input(Fore.YELLOW + "Número de serie: " + Style.RESET_ALL).strip()
            if validar_serial(serial):
                break
            print(Fore.RED + "Número de serie inválido. No se permiten espacios ni símbolos.")
        
        observaciones = input(Fore.YELLOW + "Observaciones (opcional): " + Style.RESET_ALL).strip() or "Ninguna"

        if not all([placa, tipo, marca, modelo, serial]):
            print(Fore.RED + "\n❌ Error: Todos los campos son obligatorios excepto observaciones.")
            pausar_pantalla()
            return

        print("\n" + Fore.CYAN + "--- Resumen del Nuevo Equipo ---")
        print(f"  {'Placa:'.ljust(15)} {placa}")
        print(f"  {'Tipo:'.ljust(15)} {tipo}")
        print(f"  {'Marca:'.ljust(15)} {marca}")
        print(f"  {'Modelo:'.ljust(15)} {modelo}")
        print(f"  {'Serial:'.ljust(15)} {serial}")
        print(f"  {'Observaciones:'.ljust(15)} {observaciones}")
        print("--------------------------------" + Style.RESET_ALL)

        if not confirmar_con_placa(placa):
             return

        nuevo_equipo = Equipo(placa=placa, tipo=tipo, marca=marca, modelo=modelo, serial=serial, observaciones=observaciones)
        db_manager.insert_equipo(nuevo_equipo)
        registrar_movimiento_inventario(placa, "Registro", f"Nuevo equipo registrado: {tipo} {marca} {modelo}", usuario)
        print(Fore.GREEN + f"\n✅ ¡Equipo con placa {placa} registrado exitosamente!")

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación de registro cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def gestionar_equipos(usuario: str):
    mostrar_encabezado("Gestión de Equipos", color=Fore.BLUE)
    try:
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para regresar." + Style.RESET_ALL)
        
        equipos_nuevos = db_manager.get_new_equipos()
        if equipos_nuevos:
            print(Fore.GREEN + "\n--- Equipos Nuevos (sin gestión) ---" + Style.RESET_ALL)
            for equipo in equipos_nuevos:
                print(f"  - Placa: {equipo['placa']}, Tipo: {equipo['tipo']}, Marca: {equipo['marca']} {Fore.CYAN}(New){Style.RESET_ALL}")
            print("--------------------------------------" + Style.RESET_ALL)

        while True:
            placa = input(Fore.YELLOW + "\nIngrese la placa del equipo a gestionar: " + Style.RESET_ALL).strip().upper()
            
            equipo_data = db_manager.get_equipo_by_placa(placa)

            if not equipo_data:
                print(Fore.RED + "❌ No se encontró un equipo con esa placa. Intente de nuevo.")
                continue
            
            equipo = Equipo(**equipo_data)
            menu_gestion_especifica(usuario, equipo)
            break 

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación de gestión cancelada.")
        pausar_pantalla()

def menu_gestion_especifica(usuario: str, equipo: Equipo):
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        mostrar_encabezado(f"Gestionando Equipo: {equipo.marca} {equipo.modelo} - PLACA: {equipo.placa}", color=Fore.GREEN)
        print(f"{Fore.CYAN}Estado actual:{Style.RESET_ALL} {equipo.estado}")
        if equipo.asignado_a: print(f"{Fore.CYAN}Asignado a:{Style.RESET_ALL} {equipo.asignado_a} ({equipo.email_asignado})")
        if equipo.fecha_devolucion_prestamo: print(f"{Fore.CYAN}Fecha devolución (Préstamo):{Style.RESET_ALL} {equipo.fecha_devolucion_prestamo}")
        print("-" * 80)
        
        if equipo.estado in ["En mantenimiento", "Pendiente Devolución a Proveedor", "Devuelto a Proveedor"]:
            print(Fore.YELLOW + f"⚠️  Este equipo está '{equipo.estado}'. Las acciones de gestión están limitadas.")
            opciones_limitadas = [
                "Ver Detalles del Equipo",
                "Ver Historial del Equipo (Excel)",
                "Volver al menú anterior"
            ]
            mostrar_menu(opciones_limitadas, "Opciones de Consulta")
            opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
            if opcion == '1':
                mostrar_detalles_equipo(equipo)
            elif opcion == '2':
                generar_excel_historico_equipo(usuario, equipo)
            elif opcion == '3':
                break
            else:
                print(Fore.RED + "❌ Opción no válida.")
                pausar_pantalla()
            continue
        
        opciones_gestion = [
            "Asignar/Prestar equipo", "Devolver equipo al inventario", "Registrar para mantenimiento",
            "Registrar para devolución a Proveedor", "Editar información del equipo", 
            "Eliminar equipo",
            "Ver Detalles del Equipo",
            "Ver Historial del Equipo (Excel)",
            "Volver al menú anterior"
        ]
        mostrar_menu(opciones_gestion, titulo=f"Opciones para {equipo.placa}")
        
        opcion = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()

        opciones_validas = {
            "1": "asignar", "2": "devolver", "3": "mantenimiento",
            "4": "proveedor", "5": "editar", "6": "eliminar",
            "7": "detalles", "8": "historial", "9": "volver"
        }
        accion = opciones_validas.get(opcion)

        if accion == "asignar": asignar_o_prestar_equipo(usuario, equipo)
        elif accion == "devolver": devolver_equipo(usuario, equipo)
        elif accion == "mantenimiento": registrar_mantenimiento(usuario, equipo)
        elif accion == "proveedor": registrar_devolucion_a_proveedor(usuario, equipo)
        elif accion == "editar": editar_equipo(usuario, equipo)
        elif accion == "eliminar":
            if eliminar_equipo(usuario, equipo): return 
        elif accion == "detalles":
            mostrar_detalles_equipo(equipo)
        elif accion == "historial":
            generar_excel_historico_equipo(usuario, equipo)
        elif accion == "volver": break
        else:
            print(Fore.RED + "❌ Opción no válida. Por favor, intente de nuevo.")
            pausar_pantalla()
        
        equipo_data_actualizado = db_manager.get_equipo_by_placa(equipo.placa)
        if not equipo_data_actualizado: break
        equipo = Equipo(**equipo_data_actualizado)

def mostrar_detalles_equipo(equipo: Equipo):
    """Muestra una vista detallada de la información de un equipo en la consola."""
    mostrar_encabezado(f"Detalles del Equipo: Placa {equipo.placa}", color=Fore.CYAN)
    print(f"  {'Placa:'.ljust(25)} {equipo.placa}")
    print(f"  {'Tipo:'.ljust(25)} {equipo.tipo}")
    print(f"  {'Marca:'.ljust(25)} {equipo.marca}")
    print(f"  {'Modelo:'.ljust(25)} {equipo.modelo}")
    print(f"  {'Serial:'.ljust(25)} {equipo.serial}")
    print("-" * 40)
    print(f"  {'Estado Actual:'.ljust(25)} {equipo.estado}")
    print(f"  {'Asignado a:'.ljust(25)} {equipo.asignado_a or 'N/A'}")
    print(f"  {'Email Asignado:'.ljust(25)} {equipo.email_asignado or 'N/A'}")
    print("-" * 40)
    print(f"  {'Fecha de Registro:'.ljust(25)} {equipo.fecha_registro}")
    print(f"  {'Fecha Devolución Préstamo:'.ljust(25)} {equipo.fecha_devolucion_prestamo or 'N/A'}")
    print(f"  {'Fecha Devolución Proveedor:'.ljust(25)} {equipo.fecha_devolucion_proveedor or 'N/A'}")
    print(f"  {'Motivo Devolución:'.ljust(25)} {equipo.motivo_devolucion or 'N/A'}")
    print("-" * 40)
    print(f"  {'Observaciones:'.ljust(25)} {equipo.observaciones or 'N/A'}")
    pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def asignar_o_prestar_equipo(usuario: str, equipo: Equipo):
    if equipo.estado != "Disponible":
        print(Fore.RED + f"❌ El equipo no está 'Disponible' (Estado actual: {equipo.estado}).")
        pausar_pantalla()
        return
    
    try:
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar." + Style.RESET_ALL)
        while True:
            tipo_asignacion_input = input(Fore.YELLOW + "Escriba 'A' para Asignación o 'P' para Préstamo: " + Style.RESET_ALL).strip().upper()
            if tipo_asignacion_input in ["A", "P"]:
                break
            print(Fore.RED + "Opción inválida. Intente de nuevo.")
        
        nombre_asignado = input(Fore.YELLOW + "Nombre de la persona: " + Style.RESET_ALL).strip()
        while True:
            email_asignado = input(Fore.YELLOW + "Email de la persona: " + Style.RESET_ALL).strip()
            if validar_email(email_asignado): break
            print(Fore.RED + "Email inválido. Intente de nuevo.")

        while True:
            observacion_asignacion = input(Fore.YELLOW + "Observación de la asignación/préstamo: " + Style.RESET_ALL).strip()
            if observacion_asignacion:
                break
            print(Fore.RED + "La observación es obligatoria. Intente de nuevo.")

        fecha_devolucion = None
        if tipo_asignacion_input == "P":
            tipo_movimiento = "Préstamo"
            while True:
                fecha_str = input(Fore.YELLOW + "Fecha de devolución (DD/MM/AAAA): " + Style.RESET_ALL).strip()
                if validar_formato_fecha(fecha_str):
                    fecha_devolucion = fecha_str
                    break
                print(Fore.RED + "Formato de fecha inválido. Intente de nuevo.")
        else:
            tipo_movimiento = "Asignación"

        print("\n" + Fore.CYAN + "--- Resumen de la Operación ---")
        print(f"  {'Acción:'.ljust(20)} {tipo_movimiento}")
        print(f"  {'Equipo (Placa):'.ljust(20)} {equipo.placa}")
        print(f"  {'Asignado a:'.ljust(20)} {nombre_asignado}")
        print(f"  {'Email:'.ljust(20)} {email_asignado}")
        if fecha_devolucion:
            print(f"  {'Fecha de Devolución:'.ljust(20)} {fecha_devolucion}")
        print(f"  {'Observación:'.ljust(20)} {observacion_asignacion}")
        print("--------------------------------" + Style.RESET_ALL)
        
        if not confirmar_con_placa(equipo.placa):
            return

        equipo.estado = "En préstamo" if tipo_movimiento == "Préstamo" else "Asignado"
        detalles_movimiento = f"{tipo_movimiento} a {nombre_asignado}. Obs: {observacion_asignacion}"
        if fecha_devolucion:
            detalles_movimiento += f". Devolución: {fecha_devolucion}"
        
        equipo.asignado_a = nombre_asignado
        equipo.email_asignado = email_asignado
        equipo.fecha_devolucion_prestamo = fecha_devolucion

        db_manager.update_equipo(equipo)
        registrar_movimiento_inventario(equipo.placa, "Asignación/Préstamo", detalles_movimiento, usuario)
        print(Fore.GREEN + f"\n✅ ¡Operación confirmada! Equipo {equipo.placa} ahora está '{equipo.estado}'.")

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def devolver_equipo(usuario: str, equipo: Equipo):
    if equipo.estado not in ["Asignado", "En préstamo"]:
        print(Fore.RED + "❌ El equipo no está asignado ni en préstamo.")
        pausar_pantalla()
        return

    try:
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar." + Style.RESET_ALL)
        while True:
            observacion_devolucion = input(Fore.YELLOW + "Motivo u observación de la devolución: " + Style.RESET_ALL).strip()
            if observacion_devolucion:
                break
            print(Fore.RED + "La observación es obligatoria para la devolución. Intente de nuevo.")
            
        print("\n" + Fore.CYAN + "--- Resumen de la Devolución ---")
        print(f"  {'Equipo (Placa):'.ljust(25)} {equipo.placa}")
        print(f"  {'Se retirará de:'.ljust(25)} {equipo.asignado_a or 'N/A'}")
        print(f"  {'Estado anterior:'.ljust(25)} {equipo.estado}")
        print(f"  {'Nuevo estado:'.ljust(25)} Disponible")
        print(f"  {'Observación de devolución:'.ljust(25)} {observacion_devolucion}")
        print("--------------------------------" + Style.RESET_ALL)

        if not confirmar_con_placa(equipo.placa):
            return

        detalles_previos = f"Devuelto por {equipo.asignado_a or 'N/A'}. Motivo: {observacion_devolucion}"
        equipo.estado = "Disponible"
        equipo.asignado_a = None
        equipo.email_asignado = None
        equipo.fecha_devolucion_prestamo = None
        
        db_manager.update_equipo(equipo)
        registrar_movimiento_inventario(equipo.placa, "Devolución a Inventario", detalles_previos, usuario)
        print(Fore.GREEN + f"\n✅ ¡Devolución confirmada! Equipo {equipo.placa} ahora está 'Disponible'.")

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("gestionar_equipo")
def editar_equipo(usuario: str, equipo: Equipo):
    try:
        os.system('cls' if os.name == 'nt' else 'clear')
        mostrar_encabezado(f"Editando Equipo: {equipo.placa}", color=Fore.BLUE)
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar." + Style.RESET_ALL)
        
        print(Fore.CYAN + "--- Información Actual del Equipo ---")
        print(f"  Tipo:          {equipo.tipo}")
        print(f"  Marca:         {equipo.marca}")
        print(f"  Modelo:        {equipo.modelo}")
        print(f"  Serial:        {equipo.serial}")
        print("---------------------------------------" + Style.RESET_ALL)
        print("\nDeje el campo en blanco y presione Enter para mantener el valor actual.")

        tipo_nuevo = input(Fore.YELLOW + f"Tipo ({Fore.CYAN}{equipo.tipo}{Fore.YELLOW}): " + Style.RESET_ALL).strip() or equipo.tipo
        marca_nueva = input(Fore.YELLOW + f"Marca ({Fore.CYAN}{equipo.marca}{Fore.YELLOW}): " + Style.RESET_ALL).strip() or equipo.marca
        
        while True:
            modelo_nuevo = input(Fore.YELLOW + f"Modelo ({Fore.CYAN}{equipo.modelo}{Fore.YELLOW}): " + Style.RESET_ALL).strip() or equipo.modelo
            if validar_campo_general(modelo_nuevo):
                break
            print(Fore.RED + "Modelo inválido. Solo se permiten letras, números, espacios y (- _ . ,).")
        
        while True:
            serial_nuevo = input(Fore.YELLOW + f"Serie ({Fore.CYAN}{equipo.serial}{Fore.YELLOW}): " + Style.RESET_ALL).strip() or equipo.serial
            if validar_serial(serial_nuevo):
                break
            print(Fore.RED + "Número de serie inválido. No se permiten espacios ni símbolos.")
        
        cambios = []
        if equipo.tipo != tipo_nuevo: cambios.append(f"Tipo: '{equipo.tipo}' -> '{tipo_nuevo}'")
        if equipo.marca != marca_nueva: cambios.append(f"Marca: '{equipo.marca}' -> '{marca_nueva}'")
        if equipo.modelo != modelo_nuevo: cambios.append(f"Modelo: '{equipo.modelo}' -> '{modelo_nuevo}'")
        if equipo.serial != serial_nuevo: cambios.append(f"Serial: '{equipo.serial}' -> '{serial_nuevo}'")
        
        if not cambios:
            print(Fore.YELLOW + "\nNo se detectaron cambios.")
            pausar_pantalla()
            return
            
        while True:
            motivo_edicion = input(Fore.YELLOW + "Motivo de la edición: " + Style.RESET_ALL).strip()
            if motivo_edicion:
                break
            print(Fore.RED + "El motivo de la edición es obligatorio.")

        print("\n" + Fore.CYAN + "--- Resumen de Cambios ---")
        for cambio in cambios:
            print(f"  - {cambio}")
        print(f"  - Motivo: {motivo_edicion}")
        print("--------------------------" + Style.RESET_ALL)

        if not confirmar_con_placa(equipo.placa):
            return

        equipo.tipo, equipo.marca, equipo.modelo, equipo.serial = tipo_nuevo, marca_nueva, modelo_nuevo, serial_nuevo
        db_manager.update_equipo(equipo)
        detalles_log = f"Cambios: {'; '.join(cambios)}. Motivo: {motivo_edicion}"
        registrar_movimiento_inventario(equipo.placa, "Edición", detalles_log, usuario)
        print(Fore.GREEN + f"\n✅ ¡Equipo {equipo.placa} actualizado exitosamente!")

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación de edición cancelada.")
    finally:
        pausar_pantalla()
        
@requiere_permiso("gestionar_equipo")
def registrar_mantenimiento(usuario: str, equipo: Equipo):
    try:
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar." + Style.RESET_ALL)
        
        tipos_mantenimiento = ["Preventivo", "Correctivo", "Mejora"]
        tipo_seleccionado = seleccionar_parametro(None, "Tipo de Mantenimiento", lista_opciones=tipos_mantenimiento)
        
        while True:
            observaciones_mantenimiento = input(Fore.YELLOW + "Observaciones del mantenimiento: " + Style.RESET_ALL).strip()
            if observaciones_mantenimiento:
                break
            print(Fore.RED + "Las observaciones son obligatorias.")

        print("\n" + Fore.CYAN + "--- Resumen del Mantenimiento ---")
        print(f"  {'Equipo (Placa):'.ljust(25)} {equipo.placa}")
        print(f"  {'Tipo de Mantenimiento:'.ljust(25)} {tipo_seleccionado}")
        print(f"  {'Observaciones:'.ljust(25)} {observaciones_mantenimiento}")
        print(f"  {'Nuevo estado:'.ljust(25)} En mantenimiento")
        print("-----------------------------------" + Style.RESET_ALL)
        
        if not confirmar_con_placa(equipo.placa):
            return

        equipo.estado_anterior = equipo.estado
        equipo.estado = "En mantenimiento"
        db_manager.update_equipo(equipo)
        detalles = f"Tipo: {tipo_seleccionado}. Obs: {observaciones_mantenimiento}. Estado anterior: {equipo.estado_anterior}"
        registrar_movimiento_inventario(equipo.placa, "Mantenimiento", detalles, usuario)
        print(Fore.GREEN + f"\n✅ Mantenimiento registrado. Estado cambiado a 'En mantenimiento'.")

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("devolver_a_proveedor")
def registrar_devolucion_a_proveedor(usuario: str, equipo: Equipo):
    if equipo.estado != "Disponible":
        print(Fore.RED + f"❌ El equipo debe estar 'Disponible' para ser devuelto al proveedor (Estado actual: {equipo.estado}).")
        pausar_pantalla()
        return

    try:
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar." + Style.RESET_ALL)
        
        motivos = ["Por daño", "No se necesita más", "Por hurto"]
        motivo = seleccionar_parametro(None, "Motivo de la Devolución", lista_opciones=motivos)
        
        while True:
            fecha_str = input(Fore.YELLOW + "Fecha de devolución a proveedor (DD/MM/AAAA): " + Style.RESET_ALL).strip()
            if validar_formato_fecha(fecha_str):
                fecha_devolucion = fecha_str
                break
            print(Fore.RED + "Formato de fecha inválido.")

        while True:
            observaciones = input(Fore.YELLOW + "Observaciones adicionales: " + Style.RESET_ALL).strip()
            if observaciones:
                break
            print(Fore.RED + "Las observaciones son obligatorias.")
        
        print("\n" + Fore.CYAN + "--- Resumen de Devolución a Proveedor ---")
        print(f"  {'Equipo (Placa):'.ljust(25)} {equipo.placa}")
        print(f"  {'Motivo:'.ljust(25)} {motivo}")
        print(f"  {'Fecha programada:'.ljust(25)} {fecha_devolucion}")
        print(f"  {'Observaciones:'.ljust(25)} {observaciones}")
        print(f"  {'Nuevo estado:'.ljust(25)} Pendiente Devolución a Proveedor")
        print("-----------------------------------" + Style.RESET_ALL)

        if not confirmar_con_placa(equipo.placa):
            return

        estado_anterior = equipo.estado
        equipo.estado = "Pendiente Devolución a Proveedor"
        equipo.fecha_devolucion_proveedor = fecha_devolucion
        equipo.motivo_devolucion = motivo
        equipo.observaciones = observaciones
        equipo.asignado_a = None
        equipo.email_asignado = None
        equipo.fecha_devolucion_prestamo = None

        db_manager.update_equipo(equipo)
        detalles = f"Motivo: {motivo}. Fecha prog.: {fecha_devolucion}. Obs: {observaciones}. Estado anterior: {estado_anterior}"
        registrar_movimiento_inventario(equipo.placa, "Registro Devolución Proveedor", detalles, usuario)
        print(Fore.GREEN + f"\n✅ Equipo {equipo.placa} registrado para devolución a proveedor.")

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
    finally:
        pausar_pantalla()

@requiere_permiso("eliminar_equipo")
def eliminar_equipo(usuario: str, equipo: Equipo) -> bool:
    try:
        print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para cancelar." + Style.RESET_ALL)
        
        num_movimientos = db_manager.count_movimientos_by_placa(equipo.placa)
        
        if num_movimientos > 1:
            print(Fore.RED + f"\n❌ No se puede eliminar el equipo {equipo.placa}.")
            print(Fore.YELLOW + f"   Motivo: El equipo tiene {num_movimientos} movimientos históricos registrados.")
            print(Fore.YELLOW + "   Un equipo solo puede ser eliminado si no ha tenido gestión (asignación, mantenimiento, etc.).")
            pausar_pantalla()
            return False
        
        while True:
            motivo = input(Fore.RED + "Motivo de la eliminación (obligatorio): " + Style.RESET_ALL).strip()
            if motivo:
                break
            print(Fore.RED + "El motivo no puede estar vacío.")

        print(Fore.RED + f"\n⚠️ ¿Seguro de eliminar el equipo {equipo.placa}? Esta acción es irreversible.")

        if not confirmar_con_placa(equipo.placa):
            return False

        db_manager.delete_equipo(equipo.placa)
        registrar_movimiento_inventario(equipo.placa, "Eliminación", f"Equipo eliminado. Motivo: {motivo}", usuario)
        print(Fore.GREEN + f"\n✅ Equipo {equipo.placa} eliminado.")
        pausar_pantalla()
        return True
            
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Operación cancelada.")
        pausar_pantalla()
        return False

@requiere_permiso("gestionar_pendientes")
def menu_gestionar_pendientes(usuario: str):
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        
        mantenimientos_pendientes = len([e for e in db_manager.get_all_equipos() if e.get('estado') == "En mantenimiento"])
        devoluciones_pendientes = len([e for e in db_manager.get_all_equipos() if e.get('estado') == "Pendiente Devolución a Proveedor"])

        def get_color_indicator(count):
            if count == 0: return Fore.GREEN
            elif count == 1: return Fore.YELLOW
            return Fore.RED

        color_mantenimiento = get_color_indicator(mantenimientos_pendientes)
        color_devoluciones = get_color_indicator(devoluciones_pendientes)
        
        texto_menu_mantenimiento = f"Gestionar Equipos en Mantenimiento {color_mantenimiento}({mantenimientos_pendientes}){Style.RESET_ALL}"
        texto_menu_devoluciones = f"Gestionar Devoluciones a Proveedor Pendientes {color_devoluciones}({devoluciones_pendientes}){Style.RESET_ALL}"
        
        opciones_disponibles = [texto_menu_mantenimiento, texto_menu_devoluciones, "Volver"]
        
        mostrar_menu(opciones_disponibles, titulo="Gestionar Mantenimientos y Devoluciones")
        
        opcion_input = input(Fore.YELLOW + "Seleccione una opción: " + Style.RESET_ALL).strip()
        
        opciones_map = {str(i+1): texto for i, texto in enumerate(opciones_disponibles)}
        opcion_texto = opciones_map.get(opcion_input)

        if opcion_texto and "Gestionar Equipos en Mantenimiento" in opcion_texto:
            gestionar_mantenimientos(usuario)
        elif opcion_texto and "Gestionar Devoluciones a Proveedor Pendientes" in opcion_texto:
            gestionar_devoluciones_proveedor(usuario)
        elif opcion_texto == "Volver":
            break
        else:
            print(Fore.RED + "Opción no válida.")
            pausar_pantalla()

def gestionar_mantenimientos(usuario: str):
    """Flujo mejorado para gestionar equipos en mantenimiento."""
    mostrar_encabezado("Gestionar Equipos en Mantenimiento", color=Fore.BLUE)
    try:
        while True:
            equipos_pendientes = [Equipo(**e) for e in db_manager.get_all_equipos() if e.get('estado') == "En mantenimiento"]
            if not equipos_pendientes:
                print(Fore.YELLOW + "\nNo hay equipos en mantenimiento para gestionar.")
                pausar_pantalla()
                break

            os.system('cls' if os.name == 'nt' else 'clear')
            mostrar_encabezado("Gestionar Equipos en Mantenimiento", color=Fore.BLUE)
            print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para regresar." + Style.RESET_ALL)
            print(Fore.WHITE + "\n--- Equipos en Mantenimiento ---" + Style.RESET_ALL)
            for i, equipo in enumerate(equipos_pendientes, 1):
                print(f"{Fore.YELLOW}{i}.{Style.RESET_ALL} Placa: {equipo.placa}, Tipo: {equipo.tipo}, Marca: {equipo.marca}")
            print(Fore.WHITE + "---------------------------------" + Style.RESET_ALL)

            seleccion = input(Fore.YELLOW + "\nSeleccione el equipo a gestionar: " + Style.RESET_ALL).strip()
            
            try:
                indice = int(seleccion) - 1
                if not (0 <= indice < len(equipos_pendientes)):
                    print(Fore.RED + "❌ Número no válido."); continue
                
                equipo_a_gestionar = equipos_pendientes[indice]
                
                ultimo_movimiento = db_manager.get_last_movimiento_by_placa(equipo_a_gestionar.placa)

                os.system('cls' if os.name == 'nt' else 'clear')
                mostrar_encabezado(f"Gestionando Mantenimiento: Placa {equipo_a_gestionar.placa}", color=Fore.YELLOW)
                
                print(Fore.CYAN + "--- Detalles del Equipo ---")
                print(f"  {'Placa:'.ljust(25)} {equipo_a_gestionar.placa}")
                print(f"  {'Tipo:'.ljust(25)} {equipo_a_gestionar.tipo}")
                print(f"  {'Marca:'.ljust(25)} {equipo_a_gestionar.marca}")
                print(f"  {'Modelo:'.ljust(25)} {equipo_a_gestionar.modelo}")
                print(f"  {'Serial:'.ljust(25)} {equipo_a_gestionar.serial}")

                print(Fore.CYAN + "\n--- Detalles de la Solicitud de Mantenimiento ---")
                if ultimo_movimiento and ultimo_movimiento['accion'] == 'Mantenimiento':
                    fecha_evento = datetime.strptime(ultimo_movimiento['fecha'], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
                    print(f"  {'Fecha del Evento:'.ljust(25)} {fecha_evento}")
                    print(f"  {'Usuario que Registró:'.ljust(25)} {ultimo_movimiento['usuario']}")
                    print(f"  {'Detalles:'.ljust(25)} {ultimo_movimiento['detalles']}")
                else:
                    print(Fore.YELLOW + "No se encontraron detalles específicos del registro de mantenimiento.")
                
                print(Style.RESET_ALL + "-" * 50)

                mostrar_menu(["Mantenimiento completado", "Equipo no reparable (Devolver a proveedor)"], "Acciones Disponibles")
                accion = input(Fore.YELLOW + "Seleccione una acción: " + Style.RESET_ALL).strip()

                if accion == '1':
                    while True:
                        observacion = input(Fore.YELLOW + "Observaciones de la finalización del mantenimiento: " + Style.RESET_ALL).strip()
                        if observacion:
                            break
                        print(Fore.RED + "La observación es obligatoria.")

                    nuevo_estado = equipo_a_gestionar.estado_anterior or "Disponible"
                    if not confirmar_con_placa(equipo_a_gestionar.placa): continue

                    equipo_a_gestionar.estado = nuevo_estado
                    equipo_a_gestionar.estado_anterior = None
                    db_manager.update_equipo(equipo_a_gestionar)
                    registrar_movimiento_inventario(equipo_a_gestionar.placa, "Mantenimiento Completado", f"Estado restaurado a '{nuevo_estado}'. Obs: {observacion}", usuario)
                    print(Fore.GREEN + f"\n✅ Equipo {equipo_a_gestionar.placa} ahora está '{nuevo_estado}'.")

                elif accion == '2':
                    fue_retirado = False
                    observacion_retiro = ""

                    if equipo_a_gestionar.estado_anterior in ["Asignado", "En préstamo"]:
                        print(Fore.YELLOW + f"\n⚠️ El equipo estaba asignado a '{equipo_a_gestionar.asignado_a}'.")
                        confirmacion_retiro = input(Fore.YELLOW + "¿Desea retirarlo para continuar con la devolución? (S/N): " + Style.RESET_ALL).strip().upper()
                        
                        if confirmacion_retiro == 'S':
                            while True:
                                observacion_retiro = input(Fore.YELLOW + "Observaciones del retiro del equipo al usuario: " + Style.RESET_ALL).strip()
                                if observacion_retiro:
                                    break
                                print(Fore.RED + "La observación del retiro es obligatoria.")
                            
                            fue_retirado = True
                        else:
                            print(Fore.YELLOW + "Operación cancelada. El equipo no será devuelto.")
                            pausar_pantalla()
                            continue

                    print(Fore.CYAN + "\nIniciando el registro para devolución a proveedor...")
                    motivos_devolucion = ["Por daño", "No se necesita más", "Por hurto"]
                    motivo_devolucion = seleccionar_parametro(None, "Motivo de la Devolución", lista_opciones=motivos_devolucion)

                    while True:
                        fecha_devolucion_str = input(Fore.YELLOW + "Fecha de devolución a proveedor (DD/MM/AAAA): " + Style.RESET_ALL).strip()
                        if validar_formato_fecha(fecha_devolucion_str):
                            break
                        print(Fore.RED + "Formato de fecha inválido.")

                    while True:
                        observaciones_devolucion = input(Fore.YELLOW + "Observaciones para la devolución al proveedor: " + Style.RESET_ALL).strip()
                        if observaciones_devolucion:
                            break
                        print(Fore.RED + "Las observaciones son obligatorias.")

                    os.system('cls' if os.name == 'nt' else 'clear')
                    mostrar_encabezado("Resumen de la Operación", color=Fore.GREEN)
                    print(Fore.RED + "⚠️ Esta acción es irreversible y ejecutará múltiples pasos.")

                    print(Fore.CYAN + "\n--- Detalles del Equipo ---")
                    print(f"  {'Placa:'.ljust(25)} {equipo_a_gestionar.placa}")
                    print(f"  {'Tipo:'.ljust(25)} {equipo_a_gestionar.tipo}")
                    print(f"  {'Marca:'.ljust(25)} {equipo_a_gestionar.marca}")
                    print(f"  {'Modelo:'.ljust(25)} {equipo_a_gestionar.modelo}")
                    print(f"  {'Serial:'.ljust(25)} {equipo_a_gestionar.serial}")

                    if fue_retirado:
                        print(Fore.CYAN + "\n--- Acción 1: Retiro de Equipo a Usuario ---")
                        print(f"  {'Se retirará de:'.ljust(25)} {equipo_a_gestionar.asignado_a}")
                        print(f"  {'Observación del retiro:'.ljust(25)} {observacion_retiro}")

                    print(Fore.CYAN + "\n--- Acción 2: Devolución a Proveedor ---")
                    print(f"  {'Motivo:'.ljust(25)} {motivo_devolucion}")
                    print(f"  {'Fecha Programada:'.ljust(25)} {fecha_devolucion_str}")
                    print(f"  {'Observaciones:'.ljust(25)} {observaciones_devolucion}")
                    print(f"  {'Nuevo estado final:'.ljust(25)} Pendiente Devolución a Proveedor")
                    print(Style.RESET_ALL + "-" * 50)

                    print(Fore.YELLOW + "Para confirmar TODA la operación, ingrese la placa del equipo.")
                    if not confirmar_con_placa(equipo_a_gestionar.placa):
                        print(Fore.RED + "Confirmación fallida. Operación cancelada.")
                        pausar_pantalla()
                        continue

                    if fue_retirado:
                        registrar_movimiento_inventario(equipo_a_gestionar.placa, "Devolución a Inventario", f"Retirado de {equipo_a_gestionar.asignado_a}. Motivo: {observacion_retiro}", usuario)

                    equipo_a_gestionar.estado = "Pendiente Devolución a Proveedor"
                    equipo_a_gestionar.estado_anterior = "En mantenimiento"
                    equipo_a_gestionar.asignado_a = None
                    equipo_a_gestionar.email_asignado = None
                    equipo_a_gestionar.fecha_devolucion_prestamo = None
                    equipo_a_gestionar.fecha_devolucion_proveedor = fecha_devolucion_str
                    equipo_a_gestionar.motivo_devolucion = motivo_devolucion
                    equipo_a_gestionar.observaciones = observaciones_devolucion

                    db_manager.update_equipo(equipo_a_gestionar)
                    detalles_log_devolucion = f"Motivo: {motivo_devolucion}. Fecha prog.: {fecha_devolucion_str}. Obs: {observaciones_devolucion}. Proceso iniciado desde Mantenimiento."
                    registrar_movimiento_inventario(equipo_a_gestionar.placa, "Registro Devolución Proveedor", detalles_log_devolucion, usuario)

                    print(Fore.GREEN + "\n✅ ¡Operación completada! El equipo ha sido retirado y marcado para devolución al proveedor.")

                else: 
                    print(Fore.RED + "❌ Acción no válida.")

            except ValueError:
                print(Fore.RED + "❌ Entrada inválida. Ingrese un número.")
            pausar_pantalla()
    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Regresando al menú anterior.")
        pausar_pantalla()

def gestionar_devoluciones_proveedor(usuario: str):
    """Flujo mejorado para confirmar o rechazar devoluciones a proveedor."""
    mostrar_encabezado("Gestionar Devoluciones a Proveedor", color=Fore.BLUE)
    try:
        while True:
            equipos_pendientes = [Equipo(**e) for e in db_manager.get_all_equipos() if e.get('estado') == "Pendiente Devolución a Proveedor"]
            if not equipos_pendientes:
                print(Fore.YELLOW + "\nNo hay devoluciones pendientes para gestionar.")
                pausar_pantalla()
                break

            os.system('cls' if os.name == 'nt' else 'clear')
            mostrar_encabezado("Gestionar Devoluciones a Proveedor", color=Fore.BLUE)
            print(Fore.CYAN + "💡 Puede presionar Ctrl+C en cualquier momento para regresar." + Style.RESET_ALL)
            print(Fore.WHITE + "\n--- Devoluciones Pendientes ---" + Style.RESET_ALL)
            for i, equipo in enumerate(equipos_pendientes, 1):
                print(f"{Fore.YELLOW}{i}.{Style.RESET_ALL} Placa: {equipo.placa}, Fecha Prog.: {equipo.fecha_devolucion_proveedor}, Motivo: {equipo.motivo_devolucion}")
            print(Fore.WHITE + "---------------------------------" + Style.RESET_ALL)

            seleccion = input(Fore.YELLOW + "Seleccione el equipo a gestionar: " + Style.RESET_ALL).strip()
            
            try:
                indice = int(seleccion) - 1
                if not (0 <= indice < len(equipos_pendientes)):
                    print(Fore.RED + "❌ Número no válido."); continue
                
                equipo_a_gestionar = equipos_pendientes[indice]

                ultimo_movimiento = db_manager.get_last_movimiento_by_placa(equipo_a_gestionar.placa)

                os.system('cls' if os.name == 'nt' else 'clear')
                mostrar_encabezado(f"Gestionando Devolución: Placa {equipo_a_gestionar.placa}", color=Fore.YELLOW)
                
                print(Fore.CYAN + "--- Detalles del Equipo ---")
                print(f"  {'Placa:'.ljust(25)} {equipo_a_gestionar.placa}")
                print(f"  {'Tipo:'.ljust(25)} {equipo_a_gestionar.tipo}")
                print(f"  {'Marca:'.ljust(25)} {equipo_a_gestionar.marca}")
                print(f"  {'Modelo:'.ljust(25)} {equipo_a_gestionar.modelo}")
                print(f"  {'Serial:'.ljust(25)} {equipo_a_gestionar.serial}")

                print(Fore.CYAN + "\n--- Detalles de la Solicitud de Devolución ---")
                if ultimo_movimiento and ultimo_movimiento['accion'] == 'Registro Devolución Proveedor':
                    fecha_evento = datetime.strptime(ultimo_movimiento['fecha'], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
                    print(f"  {'Fecha del Evento:'.ljust(25)} {fecha_evento}")
                    print(f"  {'Usuario que Registró:'.ljust(25)} {ultimo_movimiento['usuario']}")
                    print(f"  {'Motivo:'.ljust(25)} {equipo_a_gestionar.motivo_devolucion}")
                    print(f"  {'Fecha Programada:'.ljust(25)} {equipo_a_gestionar.fecha_devolucion_proveedor}")
                    print(f"  {'Observaciones:'.ljust(25)} {equipo_a_gestionar.observaciones}")
                else:
                    print(Fore.YELLOW + "No se encontraron detalles específicos del registro de devolución.")

                print(Style.RESET_ALL + "-" * 50)

                mostrar_menu(["Confirmar Devolución", "Rechazar Devolución", "Cancelar"], "Acciones Disponibles")
                accion = input(Fore.YELLOW + "Seleccione una acción: " + Style.RESET_ALL).strip()

                if accion == '1': # Confirmar Devolución
                    while True:
                        observacion = input(Fore.YELLOW + "Observaciones de la confirmación (ej: nro. de guía): " + Style.RESET_ALL).strip()
                        if observacion: break
                        print(Fore.RED + "La observación es obligatoria.")
                    
                    if not confirmar_con_placa(equipo_a_gestionar.placa): continue

                    equipo_a_gestionar.estado = "Devuelto a Proveedor"
                    equipo_a_gestionar.observaciones = observacion
                    db_manager.update_equipo(equipo_a_gestionar)
                    registrar_movimiento_inventario(equipo_a_gestionar.placa, "Devolución a Proveedor Completada", f"Devolución confirmada. Obs: {observacion}", usuario)
                    print(Fore.GREEN + f"\n✅ Equipo {equipo_a_gestionar.placa} marcado como 'Devuelto a Proveedor'.")

                elif accion == '2': # Rechazar Devolución
                    while True:
                        observacion = input(Fore.YELLOW + "Motivo del rechazo de la devolución: " + Style.RESET_ALL).strip()
                        if observacion: break
                        print(Fore.RED + "El motivo del rechazo es obligatorio.")
                    
                    if not confirmar_con_placa(equipo_a_gestionar.placa): continue
                    
                    equipo_a_gestionar.estado = "Disponible"
                    equipo_a_gestionar.estado_anterior = "Pendiente Devolución a Proveedor"
                    equipo_a_gestionar.observaciones = observacion
                    db_manager.update_equipo(equipo_a_gestionar)
                    registrar_movimiento_inventario(equipo_a_gestionar.placa, "Rechazo Devolución Proveedor", f"Devolución rechazada. Motivo: {observacion}", usuario)
                    print(Fore.GREEN + f"\n✅ Devolución rechazada. Equipo {equipo_a_gestionar.placa} vuelve a estar 'Disponible'.")

                elif accion == '3': # Cancelar
                    print(Fore.YELLOW + "Operación cancelada.")
                
                else:
                    print(Fore.RED + "❌ Acción no válida.")

            except ValueError:
                print(Fore.RED + "❌ Entrada inválida. Ingrese un número.")
            pausar_pantalla()

    except KeyboardInterrupt:
        print(Fore.CYAN + "\n🚫 Regresando al menú anterior.")
        pausar_pantalla()

@requiere_permiso("ver_inventario")
def ver_inventario_consola():
    mostrar_encabezado("Inventario Actual de Equipos Activos")
    inventario = db_manager.get_equipos_activos()
    if not inventario:
        print(Fore.YELLOW + "\nEl inventario activo está vacío.")
    else:
        print(f"{Fore.CYAN}{'PLACA':<12} {'TIPO':<15} {'MARCA':<15} {'MODELO':<20} {'ESTADO':<30} {'ASIGNADO A':<20}{Style.RESET_ALL}")
        print(Fore.CYAN + "="*112 + Style.RESET_ALL)
        for equipo in inventario:
            estado_color = Fore.WHITE
            if equipo['estado'] == "Disponible": estado_color = Fore.GREEN
            elif equipo['estado'] in ["Asignado", "En préstamo"]: estado_color = Fore.YELLOW
            elif equipo['estado'] == "En mantenimiento": estado_color = Fore.MAGENTA
            elif equipo['estado'] == "Pendiente Devolución a Proveedor": estado_color = Fore.LIGHTYELLOW_EX
            asignado_a = equipo.get('asignado_a') or 'N/A'
            print(f"{equipo['placa']:<12} {equipo['tipo']:<15} {equipo['marca']:<15} {equipo['modelo']:<20} {estado_color}{equipo['estado']:<30}{Style.RESET_ALL} {asignado_a:<20}")
    pausar_pantalla()